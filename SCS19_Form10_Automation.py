from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os.path, glob
import pandas as pd
import numpy as np
import os
from pathlib import Path


#Find and Sort .TXT Raw Data Files
file_name_list = []
for root,dirs,files in os.walk(data_directory_path):
    for names in files:
        if names.endswith('.txt'):
            file_name_list.append(os.path.join(root,names))
            file_name_list.sort(key=os.path.getmtime)

currentRow = startRow

def execute(data_directory_path, workbook, currentRow):

#Load Raw Data
    
    form10 = pd.read_csv(file, sep = '\t', lineterminator='\n', header = 10)
    form10_raw_extra = pd.read_csv(file, sep = '\t', lineterminator='\n', nrows = 8)

#Set Up Excel
    workbook_path = data_directory_path + r"\\" + workbook
    wb = load_workbook(workbook_path)
    ws = wb['Form10_Raw_Data']

#Clean Raw Data
    form10.drop(['\r','FOS','HMD','LME','EW','BPR','NAW','HNS','HRN','EAF','BAF','TAF','CAF'], axis = 1, inplace = True)
    
    for x in range(0,len(form10)):
        if form10['Well'][x] == 'Slope':
            break

    form10.drop(form10.tail(len(form10)-x).index, inplace = True)

#Write Raw Data to Excel
    for r in dataframe_to_rows(form10_raw_extra, index=False, header=True):
        ws['B'+currentRow].append(r)
    
    for r in dataframe_to_rows(form10, index=False, header=True):
        ws['B'+currentRow].append(r)
        
        
#QuantityMean Calculation and write
    #form10func.QuantityMean(form10)
    
#SPK Ct
    #form10func.SPK_Ct_dataframe(form10)
    
#Replicate Diff
    #form10func.ReplicateDiff(form10)
    
#Save
    wb.save(workbook)
    
for data_file in file_name_list:
    execute(r'C:\Users\efu\Desktop\Data', 'Testbook.xlsx', currentRow)
    print(file + " is written")
    currentRow += 113

print('Done')
