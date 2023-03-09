from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import shutil
from pathlib import Path
from datetime import datetime
import form10func
import QuantStudioModule_TemplateMode
import csv
import sys





#Set the parameters for the Program


        
#Data_Directory_Path = r"C:\Users\efu\Desktop\TestData"
#Excel_File_Name = 'EricFu_19Template_EmptyForm10_Macro.xlsm'
#SlopeMin = -3.7
#SlopeMax = -3.1
#LOD = 35
#R2 = .998

SlopeMin = float(sys.argv[3])
SlopeMax = float(sys.argv[4])
LOD = float(sys.argv[5])
R2 = float(sys.argv[6])

SlopeMin = float(SlopeMin)
SlopeMax = float(SlopeMax)
LOD = float(LOD)
R2 = float(R2)




def execute(Data_Directory_Path, Excel_File_Name, SlopeMin, SlopeMax, LOD, R2):


#Find and Sort .TXT Raw Data Files
        
    excel_file_exists = False
    backup_file_exists = False
    
    file_path_list = []
    
    for root,dirs,files in os.walk(Data_Directory_Path):
        for names in files:
            if names == Excel_File_Name:
                excel_file_exists = True
                
            if names.endswith('.txt') and names.startswith('MODIFIED_QUANTSTUDIO_FILE')==False:
                file_path_list.append(os.path.join(root,names))
                file_path_list.sort(key=os.path.getmtime)
            
            if names.startswith('Form10_BackUp'):
                Backup_Filename = os.path.join(root,names)
                backup_file_exists = True

    file_count = 0 


#Set Up Informational .txt Files

    Plate_Info_Exists = False
    Re_Run_Sample_Info_Exists = False

    for file in file_path_list:
        if file.split('\\')[-1] == 'Plate_Information.txt':
            Plate_Info_Exists = True
        if file.split('\\')[-1] == 'Re_Run_Sample_Info.txt':
            Re_Run_Sample_Info_Exists = True

    if Plate_Info_Exists == False: #Create Plate_Information.txt
        with open(Data_Directory_Path + '\\' + 'Plate_Information.txt', 'w', newline='') as Plate_Information:
            Plate_Information.write('Plate\tDate_Processed\tFile_Length')
        Plate_Information.close()

    Plate_Info_DF = pd.read_csv(Data_Directory_Path + '\\' + 'Plate_Information.txt', sep = '\t', lineterminator = '\r', header = 0)
    
    files_before = len(Plate_Info_DF['Plate'])
    

    #Read File into Re_Run_Sample_Info_DATAFRAME, append dataframe, save as file:
    
    if Re_Run_Sample_Info_Exists == False: #Create Re_Run_Sample_Info.txt
        with open(Data_Directory_Path + '\\' + 'Re_Run_Sample_Info.txt', 'w') as Re_Run_Sample_Info:
            Re_Run_Sample_Info.write('Need_Re_Run_Plate\tNeed_Re_Run_Samples\tReason_Re_Run\tRe_Run_Plate\tRe_Run_Samples')
        Re_Run_Sample_Info.close()

    Re_Run_Sample_Info_Dataframe = pd.read_csv(Data_Directory_Path + '\\' + 'Re_Run_Sample_Info.txt', sep = '\t', lineterminator = '\n', header = 0)    
    

    
    #Initialize variables outside the data file loop
    
    row_where_file_ends = len(Plate_Info_DF.File_Length) * 109
    currentRow = 1 + row_where_file_ends

    #healthy_samples_DF = pd.DataFrame(columns = ['Plate', 'Sample Names', 'Qty Mean'])



    #Only Recent Files get processed
            
    file_path_list = [i for i in file_path_list if i.split('\\')[-1][:-4] not in list(Plate_Info_DF['Plate'])]
    file_path_list = [i for i in file_path_list if i.split('\\')[-1] != 'Plate_Information.txt']
    file_path_list = [i for i in file_path_list if i.split('\\')[-1] != 'Re_Run_Sample_Info.txt']
        #No modified quant studio files either

    
#Set Up Excel
        
    workbook_path = Data_Directory_Path + '\\' + Excel_File_Name

    if excel_file_exists == True:
        wb = load_workbook(workbook_path, keep_vba=True)
    else:

        #SCS TEMPLATE LOCATION (Not DATA_DIRECTORY_PATH)-saves an excel to data directory
        wb_init = load_workbook(r'C:\Users\efu\Documents' + '\\' + Excel_File_Name, keep_vba=True)
        wb_init.save(workbook_path)

        wb = load_workbook(workbook_path, keep_vba=True)
        ws = wb['Form10_Raw Data']

    ws = wb['Form10_Raw Data']

    if 'Plate Report' in wb.sheetnames:
        del wb['Plate Report']

    plate_report_ws = wb.create_sheet('Plate Report')

    if 'Re Run Samples' in wb.sheetnames:
        re_run_samples_ws = wb['Re Run Samples']
    else:
        re_run_samples_ws = wb.create_sheet('Re Run Samples')


    
#Create Back-Up Project File
    
    now = datetime.now()

    if excel_file_exists == True: #Preserve Current File
        if backup_file_exists == True:
            os.rename(Backup_Filename, Data_Directory_Path + '\\' + 'Form10_BackUp_' + Excel_File_Name[:-5] + '_' + str(now.strftime('%m-%d-%Y_time_%H_%M')) + '.xlsm')
            shutil.copyfile(Data_Directory_Path + '\\' + Excel_File_Name, Data_Directory_Path + '\\' + 'Form10_BackUp_' + Excel_File_Name[:-5] + '_' + str(now.strftime('%m-%d-%Y_time_%H_%M')) + '.xlsm')
        else:
            shutil.copyfile(Data_Directory_Path + '\\' + Excel_File_Name, Data_Directory_Path + '\\' + 'Form10_BackUp_' + Excel_File_Name[:-5] + '_' + str(now.strftime('%m-%d-%Y_time_%H_%M')) + '.xlsm')
      

#Eliminate the 1 row off issue and Welcome the next batch
    
    if row_where_file_ends > 0:
        ws['A' + str(row_where_file_ends)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        ws['B' + str(row_where_file_ends)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        ws['C' + str(row_where_file_ends)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        ws['D' + str(row_where_file_ends)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        ws['E' + str(row_where_file_ends)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        ws['F' + str(row_where_file_ends)] = 'Files Processed ' + str(now.strftime('%Y-%m-%d %H:%M')) + ' and after'

#Iterate over every file
        
    for data_file in file_path_list:

                
#Test the Data file

        test = pd.read_csv(data_file, sep = '\t', skip_blank_lines = False, usecols = [0])

        if test.columns[0] == 'SDS 2.4':

            form10 = pd.read_csv(data_file, sep = '\t', lineterminator='\n', header = 10)
            form10_raw_extra = pd.read_csv(data_file, sep = '\t', lineterminator='\n', nrows = 9)

        else:

            wb.save(workbook_path)

            print('')
            print('Processing Quant Studio File: ' + data_file.split('\\')[-1])

            #Run QuantStudio function
            QuantStudioModule_TemplateMode.Run_Quant_Studio_Form10(Data_Directory_Path, data_file, Excel_File_Name, file_count, files_before, SlopeMin, SlopeMax, LOD, R2)



            #Reset Excel Variables

            file_count += 1
            
            wb = load_workbook(workbook_path, keep_vba=True)

            ws = wb['Form10_Raw Data']
            plate_report_ws = wb['Plate Report']
            re_run_samples_ws = wb['Re Run Samples']

            Re_Run_Sample_Info_Dataframe = pd.read_csv(Data_Directory_Path + '\\' + 'Re_Run_Sample_Info.txt', sep = '\t', lineterminator = '\n', header = 0)    

            Plate_Info_DF = pd.read_csv(Data_Directory_Path + '\\' + 'Plate_Information.txt', sep = '\t', lineterminator = '\r', header = 0)
            #files_before doesn't reset
            
            row_where_file_ends = len(Plate_Info_DF.File_Length) * 109

            for root,dirs,files in os.walk(Data_Directory_Path):
                for names in files:
                    if names.startswith('Form10_BackUp'):
                        Backup_Filename = os.path.join(root,names)

            continue
        
#Clean Raw Data
        
        form10.drop(['\r','FOS','HMD','LME','EW','BPR','NAW','HNS','HRN','EAF','BAF','TAF','CAF'], axis = 1, inplace = True)


#Write Raw Data to Excel

        open_data_file = open(data_file, 'r')

        currentRow = (file_count + files_before) * 109 + 1
        currentCell = None
        currentCol = 2

        while True:
        
            line = open_data_file.readline()

            if not line:
                break

            if line in ['\n', '\r', '\r\n'] or line.startswith('NAP'):
                continue

            fields = line.split("\t") #separate line by tabs into up to 34 cells
            for fieldIndex in range(len(fields)): #copy up to 21 cells in each line
            
                currentCell = ws.cell(row=currentRow, column=currentCol)

                if fields[fieldIndex]: #only try to copy field if it has a value
                    if (fields[fieldIndex][0].isdigit() or fields[fieldIndex].startswith("-")): #set cell value as a float if it begins with a number or a "-"

                        try: #copy value as a float
                            currentCell.value = float(fields[fieldIndex])

                        except: #copy value as text
                            currentCell.value = fields[fieldIndex]

                    else: #copy value as text
                        currentCell.value = fields[fieldIndex]

                currentCol += 1

                if (currentCol == 23 or fieldIndex == len(fields)-1): #move to the next row in sheet after 21 columns or at the end of the row
                    currentCol = 2
                    currentRow += 1
                    break

        open_data_file.close()

        print(data_file + " is written")

        

        #Add File Information to PlateInformation.txt; Info not contained in Plate_Info_DF however
        
        with open(Data_Directory_Path + '\\' + 'Plate_Information.txt', 'a', newline = '\r') as Plate_Information:
            Plate_Information.write('\n'+data_file.split('\\')[-1][:-4]+'\t'+str(now.strftime('%Y-%m-%d %H:%M'))+'\t'+str(len(form10)+len(form10_raw_extra)+2))
        Plate_Information.close()
        

#Form 10 Table
        
        #Manage Paste Offset

        Paste_row = row_where_file_ends + 20

        Sample_Name_List = form10func.Get_Sample_Names(form10)

        AC_exist = False
        for name in Sample_Name_List:
            if name.startswith('AC'):
                Paste_row -= 1
                AC_exist = True
                break
        

        #Paste Calculation Header
        
        ws['X' + str(Paste_row-1)] = 'Sample Name'
        
        ws['Y' + str(Paste_row-1)] = 'Qty Mean'
        
        ws['Z' + str(Paste_row-1)] = 'Replicate Difference'
        
        ws['AA' + str(Paste_row-1)] = '(+) SPK Value'
        ws['AB' + str(Paste_row-1)] = 'SPK Difference'
        ws['AC' + str(Paste_row-1)] = 'Inhibition?'
        
        #Paste Sample Names
        
        #Sample_Name_List = form10func.Get_Sample_Names(form10)
        for name in Sample_Name_List:
            ws['X' + str(Paste_row)] = name
            Paste_row += 1
        
        #QuantityMean Calculation and write
        
        Paste_row = row_where_file_ends + 20 #Reset Paste_Row position

        if AC_exist == True:
            Paste_row -= 1
         
        QuantityMeans = form10func.QtyMean_Dataframe(form10)
        for Qty_mean_value in QuantityMeans['Qty Mean']:
            ws['Y' + str(Paste_row)] = Qty_mean_value
            Paste_row += 1
    
        #Replicate Diff
        Paste_row = row_where_file_ends + 20 #Reset Paste_Row position

        if AC_exist == True:
            Paste_row -= 1
            
        Replicate_Differences = form10func.Replicate_Diff_Dataframe(form10)
        for value in Replicate_Differences['Ct diff']:
            ws['Z' + str(Paste_row)] = value
            Paste_row+=1
            
        #SPK Ct
        Paste_row = row_where_file_ends + 20 #Reset Paste_Row position

        if AC_exist == True:
            Paste_row -= 1
            
        if Sample_Name_List[0].startswith('A'):
            Paste_row += 1
        
        SPK_Ct = form10func.SPK_Ct_dataframe(form10) #SPK_Ct will be used for a list of C_Samples
        for x in range(len(SPK_Ct['SPK Values'])):
            ws['AA' + str(Paste_row)] = SPK_Ct['SPK Values'][x]
            ws['AB' + str(Paste_row)] = SPK_Ct['SPK Value Differences'][x]
            ws['AC' + str(Paste_row)] = SPK_Ct['Inhibition?'][x]
            Paste_row += 1

        #SPK Control Information
        Paste_row += 2
        ws['AA' + str(Paste_row)] = SPK_Ct['SPK Control Value'][0]
        ws['AA' + str(Paste_row)].font = Font(size = 13, bold = True)

#Style Formatting
        
        #ws['U' + str(row_where_file_ends+1)] = 1 + file_count + files_before
        #ws['U' + str(row_where_file_ends+1)].font = Font(size = 16, bold = True)
        #ws['U' + str(row_where_file_ends+1)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        #ws['U' + str(row_where_file_ends+1)].alignment = Alignment(horizontal = 'center')

        row_where_file_ends = row_where_file_ends + 109

##############################################################################################################################

        
#Plate Report

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
        
        plate_report_ws['A'+ str(Paste_row)] = 'Plate ' + str(file_count+1)
        plate_report_ws['B'+ str(Paste_row)] = 'Re-Run Status'
        plate_report_ws['C'+ str(Paste_row)] = 'Samples'
        plate_report_ws['D'+ str(Paste_row)] = 'Reason'
        plate_report_ws['F'+ str(Paste_row)] = 'Table'

        plate_report_ws['A' + str(Paste_row+1)] = data_file.split('\\')[-1][:-4]
        plate_report_ws['A'+ str(Paste_row+1)].font = Font(bold = True)
        plate_report_ws['A' + str(Paste_row+1)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')

        plate_report_ws['A'+ str(Paste_row)].font = Font(bold = True)

#Plate Report Table
        
        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
        
        plate_report_ws['F' + str(Paste_row+1)] = 'Sample Name'
        
        plate_report_ws['G' + str(Paste_row+1)] = 'Qty Mean'
        
        plate_report_ws['H' + str(Paste_row+1)] = 'Replicate Difference'
        
        plate_report_ws['I' + str(Paste_row+1)] = '(+) SPK Value'
        plate_report_ws['J' + str(Paste_row+1)] = 'SPK Difference'
        plate_report_ws['K' + str(Paste_row+1)] = 'Inhibition?'
        

        for name in Sample_Name_List:
            plate_report_ws['F' + str(Paste_row+2)] = name
            Paste_row += 1

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position

        for Qty_mean_value in QuantityMeans['Qty Mean']:
            plate_report_ws['G' + str(Paste_row+2)] = Qty_mean_value
            Paste_row += 1

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
            
        for value in Replicate_Differences['Ct diff']:
            plate_report_ws['H' + str(Paste_row+2)] = value
            Paste_row+=1

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position

        if Sample_Name_List[0].startswith('AC'):
            Paste_row += 1
        
        for x in range(len(SPK_Ct['SPK Values'])):
            plate_report_ws['I' + str(Paste_row+2)] = SPK_Ct['SPK Values'][x]
            plate_report_ws['J' + str(Paste_row+2)] = SPK_Ct['SPK Value Differences'][x]
            plate_report_ws['K' + str(Paste_row+2)] = SPK_Ct['Inhibition?'][x]
            Paste_row += 1

        Paste_row += 2
        plate_report_ws['I' + str(Paste_row+2)] = SPK_Ct['SPK Control Value'][0]
        plate_report_ws['I' + str(Paste_row+2)].font = Font(size = 12, bold = True)

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
        

#ReRun Status

        re_run_status = 'green'
        re_run_samples_replicates = 'not yet'
        re_run_samples_inhib = 'not yet'

        #Status Color Indicator Default is Green
        plate_report_ws['B'+ str(Paste_row+1)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

        #STDs 1-7
        STD_additional_comment = []
        
        for x in range(0,len(Replicate_Differences['Sample Name'])):
            if Replicate_Differences['Sample Name'][x].startswith('STD'):
                if Replicate_Differences['Sample Name'][x] != 'STD8':
                    if '.' in str(Replicate_Differences['Ct diff'][x]): #Numbers only
                        if abs(Replicate_Differences['Ct diff'][x]) > 1: #Fails
                            plate_report_ws['B'+ str(Paste_row+2)] = 'Re-run Plate'
                            plate_report_ws['D'+ str(Paste_row+2)] = Replicate_Differences['Sample Name'][x] + ' Failed' + ' '.join(STD_additional_comment)
                            STD_additional_comment.append('and' + Replicate_Differences['Sample Name'][x] + 'Failed') #STD 1-7 message
                            re_run_status = 'red'
                           
                        
        Paste_row = file_count * 40 + 1 #Reset Paste_Row position

        #Only STD8 fails plate
        for x in range(0,len(Replicate_Differences['Sample Name'])):
            if Replicate_Differences['Sample Name'][x] == 'STD8':
                if '.' in str(Replicate_Differences['Ct diff'][x]): #Numbers only
                    if abs(Replicate_Differences['Ct diff'][x]) > 1:
                        plate_report_ws['B'+ str(Paste_row+2)] = 'Re-run Plate'
                        plate_report_ws['D'+ str(Paste_row+2)] = 'STD 8 Failed' + ' '.join(STD_additional_comment)
                        re_run_status = 'red'


        #Grabbing all samples needing re run due to failed STD, can replace above 2 blocks if modified

        plate_fail_reason = []
        sample_fail_reason = []
                    
        for x in range(0,len(Replicate_Differences['Sample Name'])):
            if Replicate_Differences['Sample Name'][x].startswith('STD'):
                if '.' in str(Replicate_Differences['Ct diff'][x]):
                    if abs(Replicate_Differences['Ct diff'][x]) > 1:
                        plate_fail_reason.append(Replicate_Differences['Sample Name'][x])
                if str(Replicate_Differences['Ct diff'][x]).startswith('PROBLEM'):
                    plate_fail_reason.append(Replicate_Differences['Sample Name'][x])

                
        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
                                             

        #Slope, R2, LOD

        plate_report_ws['M' + str(Paste_row+1)] = 'Slope'
        plate_report_ws['N' + str(Paste_row+1)] = 'R^2'
        plate_report_ws['O' + str(Paste_row+1)] = 'NTC'
        plate_report_ws['P' + str(Paste_row+1)] = 'LOD'
        plate_report_ws['M' + str(Paste_row+2)] = str(form10.iloc[-8,1]) + ' cycles/log decade'
        plate_report_ws['N' + str(Paste_row+2)] = form10.iloc[-6,2][:-1]
            
        plate_report_ws['P' + str(Paste_row+2)] = LOD

        if QuantityMeans['NTC'][0] > 0:
            plate_report_ws['O' + str(Paste_row+2)] = QuantityMeans['NTC'][0]
        else:
            plate_report_ws['O' + str(Paste_row+2)] = 'Undetermined'

        if float(form10.iloc[-8,1]) < float(SlopeMin):
            plate_report_ws['B'+ str(Paste_row+2)] = 'Re-run Plate'
            plate_report_ws['D'+ str(Paste_row+3)] = 'Slope Requirement'
            plate_fail_reason.append('Slope Requirement')
            re_run_status = 'red'
        if float(form10.iloc[-8,1]) > float(SlopeMax):
            plate_report_ws['B'+ str(Paste_row+2)] = 'Re-run Plate'
            plate_report_ws['D'+ str(Paste_row+3)] = 'Slope Requirement'
            plate_fail_reason.append('Slope Requirement')
            re_run_status = 'red'
        if isinstance(form10.iloc[-6,2][:-1], str) == False:
            if float(form10.iloc[-6,2][:-1]) < R2:
                plate_report_ws['B'+ str(Paste_row+2)] = 'Re-run Plate'
                plate_report_ws['D'+ str(Paste_row+4)] = 'R^2 Requirement'
                plate_fail_reason.append('R2 Requirement')
                re_run_status = 'red'
        if QuantityMeans['NTC'][0] < float(LOD):
            plate_report_ws['B'+ str(Paste_row+2)] = 'Re-run Plate'
            plate_report_ws['D'+ str(Paste_row+5)] = 'NTC vs LOD Requirement'
            plate_fail_reason.append('NTC vs LOD Requirement')
            re_run_status = 'red'

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position

############################################################################################

#Re_Run_Samples Report

    
    #Connect Samples that were re_run with their former selves.
        for C_name in SPK_Ct['Sample Names']:
            C_name_with_no_plus_SPK = C_name.split('+')[0]
            C_num = C_name.split('_')[0]
            for Re_Run_Sample_Info_Index in range(0,len(Re_Run_Sample_Info_Dataframe['Need_Re_Run_Samples'])):
                if C_num == Re_Run_Sample_Info_Dataframe['Need_Re_Run_Samples'][Re_Run_Sample_Info_Index].split('_')[0]:
                    

                    #Robust dumb method FOR PASTING RE RAN SAMPLES into DF, relies on date sort for accuracy
                    Re_Run_Sample_Info_Dataframe.iloc[Re_Run_Sample_Info_Index, 3] = data_file.split('\\')[-1][:-4]
                    Re_Run_Sample_Info_Dataframe.iloc[Re_Run_Sample_Info_Index, 4] = C_name_with_no_plus_SPK


                    
    #Samples needing re-run from Replicate Difference 

        replicate_diff_issue_samples = []
        Reporting_Samples_with_replicate_diff_issue = []

        for Replicate_value_index in range(0,len(Replicate_Differences)):
            if str(Replicate_Differences['Ct diff'][Replicate_value_index]).startswith('PROBLEM'):
                    
                Reporting_Samples_with_replicate_diff_issue.append(Replicate_Differences['Sample Name'][Replicate_value_index])
                plate_report_ws['C' + str(Paste_row+6)] = ', '.join(Reporting_Samples_with_replicate_diff_issue) #paste into 1 cell
                plate_report_ws['D'+ str(Paste_row+6)] = 'Replicate Difference'

                if Replicate_Differences['Sample Name'][Replicate_value_index].startswith('C'):
                    replicate_diff_issue_samples.append(Replicate_Differences['Sample Name'][Replicate_value_index])
                
                re_run_samples_replicates = 'true'
                re_run_status = 'red'
                Paste_row += 1

                if Replicate_Differences['Sample Name'][Replicate_value_index].startswith('AC'):
                    plate_fail_reason.append('AC diff')
                                       
                #Update Re_Run_Samples with replicate diff fails
                if len(plate_fail_reason) == 0:
                    if str(Replicate_Differences['Sample Name'][Replicate_value_index].startswith('S')) == False and str(Replicate_Differences['Sample Name'][Replicate_value_index].startswith('N')) == False and str(Replicate_Differences['Sample Name'][Replicate_value_index].startswith('A')) == False:
                        Re_Run_Sample_Info_Dataframe = Re_Run_Sample_Info_Dataframe.append({'Need_Re_Run_Plate':data_file.split('\\')[-1][:-4],'Need_Re_Run_Samples':Replicate_Differences['Sample Name'][Replicate_value_index].split('+')[0],'Reason_Re_Run':'Replicate Difference'}, ignore_index=True)


    #Samples needing re-run from Plate Failiures
        if len(plate_fail_reason) > 0:
            for i in range(0,len(SPK_Ct)):
                Re_Run_Sample_Info_Dataframe = Re_Run_Sample_Info_Dataframe.append({'Need_Re_Run_Plate':data_file.split('\\')[-1][:-4],'Need_Re_Run_Samples':SPK_Ct['Sample Names'][i].split('+')[0],'Reason_Re_Run': ', '.join(plate_fail_reason) + ' Failed'}, ignore_index=True)

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
        
    #Samples needing re-run from Inhibition

        inhibited_samples = []
        
        for Inhibited_Sample_Index in range(0,len(SPK_Ct['Inhibition?'])):        

            if SPK_Ct['Inhibition?'][Inhibited_Sample_Index] == 'True':

                plate_report_ws['C' + str(Paste_row+7)] = SPK_Ct['Sample Names'][Inhibited_Sample_Index]
                plate_report_ws['D'+ str(Paste_row+7)] = 'Inhibited'
                
                if SPK_Ct['Sample Names'][Inhibited_Sample_Index].startswith('C'):
                    inhibited_samples.append(SPK_Ct['Inhibition?'][Inhibited_Sample_Index])
                
                re_run_samples_inhib = 'true'
                re_run_status = 'red'
                Paste_row += 1


                #Update Re_Run_Samples with inhibited
                if len(plate_fail_reason) == 0:
                    Re_Run_Sample_Info_Dataframe = Re_Run_Sample_Info_Dataframe.append({'Need_Re_Run_Plate':data_file.split('\\')[-1][:-4],'Need_Re_Run_Samples':SPK_Ct['Sample Names'][Inhibited_Sample_Index].split('+')[0],'Reason_Re_Run':'Inhibited'}, ignore_index=True)

       
        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
        

        if re_run_samples_replicates == 'true':
                plate_report_ws['B'+ str(Paste_row+6)] = 'Re-run Samples'

        if re_run_samples_inhib == 'true' and re_run_samples_replicates == 'not yet':
                plate_report_ws['B'+ str(Paste_row+7)] = 'Re-run Samples' #Diff Row

        if re_run_status == 'red':
                plate_report_ws['B'+ str(Paste_row+1)].fill = PatternFill(start_color='00FF0000',end_color='00FF0000',fill_type='solid')

        for x in Sample_Name_List[-9:]:
                if x.startswith('S') == False:
                        plate_report_ws['D' + str(Paste_row+1)] = 'Re-Run with a STD removed'


#Listing healthy samples
                        
        healthy_samples = []
        healthy_samples_DF = pd.DataFrame(columns = ['Plate', 'Sample Names', 'Qty Mean'])

        
        if re_run_samples_replicates == 'not yet' and re_run_samples_inhib == 'not yet' and re_run_status == 'green':

            #Healthy plate, all samples
            for i in range(len(QuantityMeans['Sample Names'])):
                if QuantityMeans['Sample Names'][i].startswith('C'):
                    
                    healthy_samples.append(QuantityMeans['Sample Names'][i])
                    healthy_samples_DF.append({'Plate':data_file.split('\\')[-1][:-4], 'Sample Names': QuantityMeans['Sample Names'][i], 'Qty Mean': QuantityMeans['Qty Mean'][i]}, ignore_index=True)
        else:
            
            for sample_name in healthy_samples:
                if sample_name in inhibited_samples or sample_name in replicate_diff_issue_samples: 
                    healthy_samples.remove(sample_name)

                else:
                    healthy_samples_DF.append({'Plate':data_file.split('\\')[-1][:-4],'Sample Names': sample_name, 'Qty Mean': QuantityMeans.loc[QuantityMeans['Sample Names'] == sample_name, ['Qty Mean']]})


        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
                                           
        for i in range(len(healthy_samples_DF)):
            plate_report_ws['R' + str(Paste_row+1)] = healthy_samples_DF['Sample Names'][i]
            plate_report_ws['S' + str(Paste_row+1)] = healthy_samples_DF['Qty Mean'][i]
            Paste_row += 1
                                           

                                
#Locate Area to write next set of Calcultions
        file_count += 1
        currentRow = (file_count + files_before) * 109 + 1

#Save Re Run Sample Info
                        
        Re_Run_Sample_Info_Dataframe.to_csv(Data_Directory_Path + '\\' + 'Re_Run_Sample_Info.txt', sep='\t', index = False, header = True)

    if files_before == 0:
        for r in dataframe_to_rows(Re_Run_Sample_Info_Dataframe, index=False, header=True):
            re_run_samples_ws.append(r)
    else:
        
        del wb['Re Run Samples']
        re_run_samples_ws = wb.create_sheet('Re Run Samples')
        
        for r in dataframe_to_rows(Re_Run_Sample_Info_Dataframe, index=False, header=True):
            re_run_samples_ws.append(r)
            


#Set Column Width for Readability

    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['X'].width = 17

    plate_report_ws.column_dimensions['A'].width = 11
    plate_report_ws.column_dimensions['B'].width = 14
    plate_report_ws.column_dimensions['C'].width = 21
    plate_report_ws.column_dimensions['D'].width = 22
    plate_report_ws.column_dimensions['F'].width = 18
    #plate_report_ws.column_dimensions['G'].width = 8
    plate_report_ws.column_dimensions['H'].width = 17
    plate_report_ws.column_dimensions['I'].width = 12
    plate_report_ws.column_dimensions['J'].width = 13
    plate_report_ws.column_dimensions['K'].width = 10
    plate_report_ws.column_dimensions['M'].width = 10
    plate_report_ws.column_dimensions['N'].width = 10
    plate_report_ws.column_dimensions['O'].width = 13
    plate_report_ws.column_dimensions['P'].width = 5


    re_run_samples_ws.column_dimensions['A'].width = 18
    re_run_samples_ws.column_dimensions['B'].width = 25
    re_run_samples_ws.column_dimensions['C'].width = 30
    re_run_samples_ws.column_dimensions['D'].width = 14
    re_run_samples_ws.column_dimensions['E'].width = 25
    
        
#Save!
    
    wb.save(workbook_path)


    #Save Back Up File
    
    if excel_file_exists == True: #Preserve Current File
        if backup_file_exists == True:
            os.rename(Backup_Filename, Data_Directory_Path + '\\' + 'Form10_BackUp_' + Excel_File_Name[:-5] + '_' + str(now.strftime('%m-%d-%Y_time_%H_%M')) + '.xlsm')
            shutil.copyfile(Data_Directory_Path + '\\' + Excel_File_Name, Data_Directory_Path + '\\' + 'Form10_BackUp_' + Excel_File_Name[:-5] + '_' + str(now.strftime('%m-%d-%Y_time_%H_%M')) + '.xlsm')
        else:
            shutil.copyfile(Data_Directory_Path + '\\' + Excel_File_Name, Data_Directory_Path + '\\' + 'Form10_BackUp_' + Excel_File_Name[:-5] + '_' + str(now.strftime('%m-%d-%Y_time_%H_%M')) + '.xlsm')


    
#Run the Program

#CMD Line: python C:\Users\efu\Desktop\Python\SCS19AutoForm10_Excelver.py "1" "2" "3" "4" "5" "6"

if __name__ == '__main__':
    execute(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6])

    #Data_Directory_Path, Excel_File_Name, Slope_Min, Slope_Max, LOD, R2
    
print('Done')




