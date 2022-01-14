#For quickly analyzing qPCR files
#Developed by Ross Fu 07/27/21 - 11/20/21
#ericrossfu@yahoo.com

import os
import sys
import csv
import time
import math
import shutil
import warnings
import openpyxl
import statistics
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


warnings.filterwarnings('ignore')   #Program Throws Warnings but does not carry out those warnings:
                                    #Data Validation Extension will be removed
                                    #Image in Y Intercept Report will be removed 


execute_style = 'VBA'
cause_error = False




#Set the parameters for the Program

if execute_style == 'Python': #Python execution cannot exclude plates yet
    
    Project_ID = 'TestData'
    Data_Directory_Path = r'O:\5 GlobalShare\Python Scripts\AutoForm10\TestFolder\TestData'
    Excel_File_Name = 'Ross_SCS_19.xlsm'
    Project_Output_Path = r'O:\5 GlobalShare\Python Scripts\AutoForm10\TestFolder\Working Directories\TestData'
    Excel_Template_Path = r'O:\5 GlobalShare\Python Scripts\AutoForm10\SCS 19.02 Master Template.xlsm'
    SlopeMin = -3.7
    SlopeMax = -3.1
    LOD = 2.5
    LOQ = 25
    R2 = .99
    
    Inhib_Diff_Criteria = 1
    Repl_Diff_Criteria = 1
    Repl_Diff_Criteria_for_STD_1_to_7 = 1
    Repl_Diff_Criteria_for_STD_8 = 1

    SCS_Path_that_ran_program = r'O:\5 GlobalShare\Python Scripts\AutoForm10\SCS 19.02 Master Template.xlsm'

    

if execute_style == 'VBA':
    
    Project_ID = sys.argv[1]
    Data_Directory_Path = sys.argv[2]
    Excel_File_Name = sys.argv[3]
    Project_Output_Path = sys.argv[4]
    Excel_Template_Path = sys.argv[5]
    SlopeMin = float(sys.argv[6])
    SlopeMax = float(sys.argv[7])
    LOQ = float(sys.argv[8])
    LOD = float(sys.argv[9])
    R2 = float(sys.argv[10])
    Inhib_Diff_Criteria = float(sys.argv[11])
    Repl_Diff_Criteria = float(sys.argv[12])
    Repl_Diff_Criteria_for_STD_1_to_7 = float(sys.argv[13])
    Repl_Diff_Criteria_for_STD_8 = float(sys.argv[14])
    SCS_Path_that_ran_program = sys.argv[15]
    
    SlopeMin = float(SlopeMin)
    SlopeMax = float(SlopeMax)
    LOQ = float(LOQ)
    LOD = float(LOD)
    R2 = float(R2)
    Inhib_Diff_Criteria = float(Inhib_Diff_Criteria)
    Repl_Diff_Criteria = float(Repl_Diff_Criteria)
    Repl_Diff_Criteria_for_STD_1_to_7 = float(Repl_Diff_Criteria_for_STD_1_to_7)
    Repl_Diff_Criteria_for_STD_8 = float(Repl_Diff_Criteria_for_STD_8)



    
now = datetime.now()
Runtime = now.strftime('%m_%d_%Y__time_%H_%M')



#############################################################################################################
###################################     Table Of Contents    ################################################
#############################################################################################################
#                                                                                                           #
#   Set the Parameters                                                                                      #
#   Table Of Contents                                                                                       #
#   Form10 Functions                                                                                        #
#   System Validation                                                                                       #
#   Find and Sort Raw Data Files (.txt)                                                                     #
#   Set Up Informational Files (.txt)                                                                       #
#   Initialize variables outside the data file loop                                                         #
#   Set Up Excel                                                                                            #
#   Eliminate the 1 row off issue and Welcome the next batch                                                #
#   Iterate over every file                                                                                 #
#   Classify the Data file                                                                                  #
#                                                                                                           #
#                                                                                                           #
#   7900:                                                                                                   #
#       Clean Raw Data                                                                                      #
#       Write Raw Data to Excel                                                                             #
#       More Cleaning: Drop the bottom, set slope, R2, & Drop NaN samples                                   #
#                                                                                                           #
#   QS:                                                                                                     #
#       Find the actual data                                                                                #
#       Process File                                                                                        #
#       Create Modified Quant Studio File                                                                   #
#       Write Data to Excel                                                                                 #
#       More Cleaning: Drop NaNs                                                                            #
#                                                                                                           #
#   7500:                                                                                                   #
#       Find the actual data                                                                                #
#       Read File                                                                                           #
#       Create Modified 7500 File                                                                           #
#       Write Data to Excel                                                                                 #
#       More Cleaning: Drop NaNs                                                                            #
#                                                                                                           #
#                                                                                                           #
#   Form 10 Table                                                                                           #
#   Style Formatting                                                                                        #
#   Plate Report                                                                                            #
#   Plate Report Table                                                                                      #
#   ReRun Status (Plates)                                                                                   #
#   Re_Run_Samples Report                                                                                   #
#   AC Report                                                                                               #
#                                                                                                           #
#############################################################################################################
#                                                                                                           #
#   Write Re Run Samples                                                                                    #
#   Set Column Width for Readability                                                                        #
#   Save!                                                                                                   #
#   Run the Program                                                                                         #
#                                                                                                           #
#############################################################################################################
#############################################################################################################
#############################################################################################################








######################################################################    FORM 10 FUNCTIONS

def Get_Sample_Names(form10):
    Sample_Name_List = []
    for x in range(1,len(form10)):
        if (form10['Sample Name'][x]) == (form10['Sample Name'][x-1]):
            Sample_Name_List.append(form10['Sample Name'][x])
    return Sample_Name_List


def QtyMean_Dataframe(form10):
    Qty_Means = pd.DataFrame()
    QtyMean_List = []
    Sample_names_list = []
    NTC = 'Undetermined'
    for x in range(1,len(form10)):
        if (form10['Sample Name'][x]) == (form10['Sample Name'][x-1]) and str(form10['Sample Name'][x].lower()).startswith('spk') == False and str(form10['Sample Name'][x]).startswith('N') == False:
            if math.isnan(form10['Quantity'][x]) and math.isnan(form10['Quantity'][x-1]) == False:
                Sample_names_list.append(form10['Sample Name'][x])
                QtyMean_List.append(float(form10['Quantity'][x-1]))
            elif math.isnan(form10['Quantity'][x-1]) and math.isnan(form10['Quantity'][x]) == False:
                Sample_names_list.append(form10['Sample Name'][x])
                QtyMean_List.append(float(form10['Quantity'][x]))
            else:
                Sample_names_list.append(form10['Sample Name'][x])
                QtyMean_List.append((float(form10['Quantity'][x])+float(form10['Quantity'][x-1]))/ 2)
        if (form10['Sample Name'][x]) == 'NTC' and (form10['Sample Name'][x-1]) == 'NTC':
            if form10['Quantity'][x] == 'Undetermined' or math.isnan(form10['Quantity'][x]):
                NTC = 'Undetermined'
            else:
                NTC = ((form10['Quantity'][x]+form10['Quantity'][x-1])/2)
    Qty_Means['Sample Name'] = Sample_names_list
    Qty_Means['Qty Mean'] = QtyMean_List
    if NTC == 'Undetermined':
        Qty_Means['NTC'] = pd.Series('Undetermined', dtype=object)
    else:
        Qty_Means['NTC'] = pd.Series(NTC, dtype = float)
    return Qty_Means


def Replicate_Diff_Dataframe(form10, LOQ):
    Replicate_Diff_Dataframe = pd.DataFrame()
    Ct_diff_list = []
    Sample_list = []
    for x in range(1,len(form10)):  
        if (form10['Sample Name'][x]) == (form10['Sample Name'][x-1]):
            Sample_list.append(form10['Sample Name'][x])  
            if (str(form10['Ct'][x]).isalpha()) != (str(form10['Ct'][x-1]).isalpha()):
            #Diff type and above AND below LOQ
                if (form10['Quantity'][x] > LOQ) != (form10['Quantity'][x-1] > LOQ):
                    Ct_diff_list.append('PROBLEM: Kind of, 1 Undet & 1 Ct but the 2 Qty values are above AND below LOQ')
                    continue
            #Diff type and above LOQ
                if form10['Quantity'][x] > LOQ or form10['Quantity'][x-1] > LOQ:
                    Ct_diff_list.append('PROBLEM: 1 Undetermined & 1 Ct and Quantity > LOQ')
            #Diff type and below LOQ
                if form10['Quantity'][x] <= LOQ or form10['Quantity'][x-1] <= LOQ:
                    Ct_diff_list.append('1 Undetermined and 1 Ct but Quantity < LOQ')                
            #2 undetermined
            elif str(form10['Ct'][x]).isalpha() == True and str(form10['Ct'][x-1]).isalpha() == True:
                Ct_diff_list.append('Undetermined')
            #Large diff and above LOQ 
            elif form10['Sample Name'][x] == 'STD8' and abs(float(form10['Ct'][x]) - float(form10['Ct'][x-1])) > Repl_Diff_Criteria_for_STD_8:
                Ct_diff_list.append('PROBLEM: LARGE DIFFERENCE ' + str(float(form10['Ct'][x]) - float(form10['Ct'][x-1])))                          #STD 8 
            elif form10['Sample Name'][x].startswith('STD') and form10['Sample Name'][x] != 'STD8' and abs(float(form10['Ct'][x]) - float(form10['Ct'][x-1])) > Repl_Diff_Criteria_for_STD_1_to_7:
                Ct_diff_list.append('PROBLEM: LARGE DIFFERENCE ' + str(float(form10['Ct'][x]) - float(form10['Ct'][x-1])))                          #STD 1-7
            elif abs(float(form10['Ct'][x]) - float(form10['Ct'][x-1])) > Repl_Diff_Criteria and (((form10['Quantity'][x] + form10['Quantity'][x-1])/2) > LOQ):
                Ct_diff_list.append('PROBLEM: LARGE DIFFERENCE ' + str(float(form10['Ct'][x]) - float(form10['Ct'][x-1])) + ' (Qty Mean > LOQ)')    #Replicates 
            elif (form10['Quantity'][x] > LOQ) != (form10['Quantity'][x-1] > LOQ):
                Ct_diff_list.append('The qty values are above AND below LOQ, but the difference is: ' + str(abs((float(form10['Ct'][x]) - float(form10['Ct'][x-1])))))
            else:
                Ct_diff_list.append(abs((float(form10['Ct'][x]) - float(form10['Ct'][x-1]))))
                
    Replicate_Diff_Dataframe['Sample Name'] = Sample_list
    Replicate_Diff_Dataframe['Ct diff'] = pd.Series(Ct_diff_list)      
    return Replicate_Diff_Dataframe


def SPK_Ct_dataframe(form10):
    SPK_Cts = pd.DataFrame()
    SPK_names = []
    SPK_values = []
    SPK_value_differences = []
    Inhibition_true_false = []
    SPK_control_values = []
    #SPK Control Average
    for x in range(1, len(form10)):
        if form10['Sample Name'][x] == 'SPK CONTROL' and form10['Sample Name'][x-1] == 'SPK CONTROL':
            if form10['Ct'][x] == 'Undetermined' or form10['Ct'][x-1] == 'Undetermined':
                SPK_avg = 'Undetermined'
            else:
                SPK_control_values = list(SPK_control_values)
                SPK_control_values.append(form10['Ct'][x-1])
                SPK_control_values.append(form10['Ct'][x])
                SPK_control_values = pd.to_numeric(SPK_control_values)
                SPK_avg = sum(SPK_control_values) / len(SPK_control_values)
    #Extracting +SPK Information
    for x in range(0,len(form10)): 
        if form10['Sample Name'][x].lower().endswith('spk'):
            SPK_names.append(form10['Sample Name'][x])
            SPK_values.append(form10['Ct'][x])
            if SPK_avg == 'Undetermined':
                SPK_Cts['SPK Value Differences'] = 'Undetermined'
                Inhibition_true_false.append('True')
            elif form10['Ct'][x] == 'Undetermined':
                Inhibition_true_false.append('True')
                SPK_value_differences.append(Inhib_Diff_Criteria + 1)
            else:
                SPK_value_difference_calculation = float(form10['Ct'][x]) - SPK_avg     #Inhibition is when +spk Ct is 1 or more greater than SPK avg
                SPK_value_differences.append(SPK_value_difference_calculation)          #The '1' is now a changeable parameter: 'Inhib_Diff_Criteria'
                if SPK_value_difference_calculation >= Inhib_Diff_Criteria:
                    Inhibition_true_false.append('True')   
                else:
                    Inhibition_true_false.append('False') 
    #Displaying Results in DataFrame
    SPK_Cts['Sample Name'] = SPK_names
    SPK_Cts['SPK Values'] = SPK_values
    SPK_Cts['SPK Control Value'] = pd.Series(SPK_avg, dtype = float)
    SPK_Cts['SPK Control Cts'] = pd.Series(SPK_control_values, dtype = float)
    SPK_Cts['Inhibition?'] = pd.Series(Inhibition_true_false, dtype = object)
    if SPK_avg != 'Undetermined':
        SPK_Cts['SPK Value Differences'] = pd.Series(SPK_value_differences, dtype = float)
    return SPK_Cts


def AC_report(form10): #Returns the 2 AC's in a plate
    AC_report_df = pd.DataFrame()
    AC_values_exist = False
    for i in range(1, len(form10)):
        if form10['Sample Name'][i] == 'AC' and form10['Sample Name'][i-1] == 'AC':
            AC_values_exist = True
            AC_report_df = AC_report_df.append({'AC_Value_Exists': True, 'Ct1':form10['Ct'][i-1], 'Ct2':form10['Ct'][i], 'Qty1':form10['Quantity'][i-1], 'Qty2':form10['Quantity'][i], 'QtyMean': ((form10['Quantity'][i-1]+form10['Quantity'][i])/2)}, ignore_index = True)
    if AC_values_exist == False:
        AC_report_df = AC_report_df.append({'AC_Value_Exists': False, 'Ct1':form10['Ct'][i-1], 'Ct2':form10['Ct'][i], 'Qty1':'NaN', 'Qty2':'NaN', 'QtyMean': ((form10['Quantity'][i-1]+form10['Quantity'][i])/2)}, ignore_index = True)
    return AC_report_df
    

def healthy_samples_DF(form10):
    return 'Ask Ross to develop this function'

def STD_info_DF(form10):

    std_df = pd.DataFrame()
    std_range = []
    std_ct1 = []
    std_ct2 = []
    statement = ''
    
    for sample in range(1, len(form10)):
        if form10['Sample Name'][sample].startswith('STD') and form10['Sample Name'][sample] == form10['Sample Name'][sample-1]:
            std_range.append(form10['Quantity'][sample])
            std_ct1.append(form10['Ct'][sample-1])
            std_ct2.append(form10['Ct'][sample])
            
    for i in range(len(std_range)):                 #counts over each standard point
        tmp_x = []                                  #a tmp list that will hold the x values for slope. The x value will be the log base 10 of the quantity (see below)
        tmp_y = []                                  #a tmp list that will hold the y values
        for x in range(len(std_range)):             #determines which point will be removed
            if x!=i:
                tmp_x.append(np.log10(std_range[x]))
                tmp_x.append(np.log10(std_range[x]))
                tmp_y.append(std_ct1[x])
                tmp_y.append(std_ct2[x])
                
        if 'Undetermined' not in std_ct1 and 'Undetermined' not in std_ct2: #normal
            tmp_y1 = []
            for y in tmp_y:
                tmp_y1.append(float(y))        
            slope, intercept = np.polyfit(tmp_x, tmp_y1, 1)    
            if SlopeMin <= slope <= SlopeMax:
                statement = "Remove STD " + str(i+1) + ": " + str(slope)[:5]
        else:
            for i in range(len(std_ct1)):
                if std_ct1[i] == 'Undetermined':
                    statement = 'STD ' + str(i+1) + ' is undetermined'
            for i in range(len(std_ct2)):
                if std_ct1[i] == 'Undetermined':
                    statement = 'STD ' + str(i+2) + ' is undetermined'
    std_df['std_range'] = std_range
    std_df['ct1'] = std_ct1
    std_df['ct2'] = std_ct2
    std_df['statement'] = statement
    return std_df


def run_log(message):
    with open(Project_Output_Path + '\\' + Project_ID + '_' + Runtime + '_AF10_run_log.txt', 'a') as z:
        z.write(message)


############################################################################################################################################################################






def execute(Project_ID, Data_Directory_Path, Excel_File_Name, Project_Output_Path, Excel_Template_Path, SlopeMin, SlopeMax, LOQ, LOD, R2, Inhib_Diff_Criteria, Repl_Diff_Criteria, Repl_Diff_Criteria_for_STD_1_to_7, Repl_Diff_Criteria_for_STD_8, SCS_Path_that_ran_program):



    
#Clean Parameters
    
    if execute_style == 'VBA':
        Project_ID = sys.argv[1]
        Data_Directory_Path = sys.argv[2]
        Excel_File_Name = sys.argv[3]
        Project_Output_Path = sys.argv[4]
        Excel_Template_Path = sys.argv[5]
        SlopeMin = float(sys.argv[6])
        SlopeMax = float(sys.argv[7])
        LOQ = float(sys.argv[8])
        LOD = float(sys.argv[9])
        R2 = float(sys.argv[10])
        Inhib_Diff_Criteria = float(sys.argv[11])
        Repl_Diff_Criteria = float(sys.argv[12])
        Repl_Diff_Criteria_for_STD_1_to_7 = float(sys.argv[13])
        Repl_Diff_Criteria_for_STD_8 = float(sys.argv[14])
        
        SlopeMin = float(SlopeMin)
        SlopeMax = float(SlopeMax)
        LOQ = float(LOQ)
        LOD = float(LOD)
        R2 = float(R2)
        Inhib_Diff_Criteria = float(Inhib_Diff_Criteria)
        Repl_Diff_Criteria = float(Repl_Diff_Criteria)
        Repl_Diff_Criteria_for_STD_1_to_7 = float(Repl_Diff_Criteria_for_STD_1_to_7)
        Repl_Diff_Criteria_for_STD_8 = float(Repl_Diff_Criteria_for_STD_8)


   
#Run Logs
    

    #Create Project_Output_Path Log
    if os.path.exists(Project_Output_Path) == False: #Project_Output_Path
        os.mkdir(Project_Output_Path)

    with open(Project_Output_Path + '\\' + Project_ID + '_' + Runtime + '_AF10_run_log.txt', 'w') as z:
        z.write('User: ' + os.environ['USERNAME'] + '\n')
        z.write('Computer Name: ' + os.environ['COMPUTERNAME'] + '\n\n')


    #Dispose of the VBS Script
    if os.path.exists(Project_Output_Path + '\\' + 'VBS_for_execute_AutoForm10.vbs'):
        os.remove(Project_Output_Path + '\\' + 'VBS_for_execute_AutoForm10.vbs')


    
#System Validation
    
    #Parameters
    if os.path.exists(Data_Directory_Path) == False:
        print('Data_Directory_Path: (' + Data_Directory_Path + ') does not exist, cant run program')
        run_log('Data_Directory_Path: (' + Data_Directory_Path + ') does not exist, cant run program')
        print('')
        input("Press Enter to close the program window")
        os._exit(0)


    #Create Data_Directory_Path Run Log
    if os.path.exists(Data_Directory_Path + '\\Audit_Logs') == False:   #Audit Logs
        os.mkdir(Data_Directory_Path + '\\Audit_Logs')                  #Audit Log copied at end of run

        
    if os.path.exists(Excel_Template_Path) == False:
        print('Excel_Template_Path: (' + Excel_Template_Path + ') does not exist, cant run program')
        run_log('Excel_Template_Path: (' + Excel_Template_Path + ') does not exist, cant run program')
        print('')
        input("Press Enter to close the program window")
        os._exit(0)


        
    #Python
    import platform
    import pkg_resources
    
    
    if platform.system() != 'Windows':
        print('This program is only validated on Windows OS, You are using ' + str(platform.system()))
        run_log('This program is only validated on Windows OS, You are using ' + str(platform.system()))
        print('')
        input("Press Enter to close the program window")
        os._exit(0)
        
    if platform.python_version() != '3.9.6':
        print('This program is only validated on Python ver 3.9.6, You are using ' + str(platform.python_version()))
        run_log('This program is only validated on Python ver 3.9.6, You are using ' + str(platform.python_version()))
        print('')
        input("Press Enter to close the program window")
        os._exit(0)
        
    if np.__version__ != '1.21.1':
        print('This program is only validated on Numpy ver 1.21.1, You are using ' + str(numpy.__version__))
        run_log('This program is only validated on Numpy ver 1.21.1, You are using ' + str(numpy.__version__))
        print('')
        input("Press Enter to close the program window")
        os._exit(0)
        
    if pd.__version__ != '1.3.1':
        print('This program is only validated on Pandas ver 1.3.1, You are using pandas ' + str(pd.__version__))
        run_log('This program is only validated on Pandas ver 1.3.1, You are using ' + str(pd.__version__))
        print('')
        input("Press Enter to close the program window")
        os._exit(0)
        
    if openpyxl.__version__ != '3.0.7':
        print('This program is only validated on Openpyxl version 3.0.7, You are using ' + str(openpyxl.__version__))
        run_log('This program is only validated on Openpyxl version 3.0.7, You are using ' + str(openpyxl.__version__))
        print('')
        input("Press Enter to close the program window")
        os._exit(0)

    if pkg_resources.get_distribution("DateTime").version != '4.3':
        print('This program is only validated on DateTime version 4.3, You are using ' + str(pkg_resources.get_distribution("DateTime").version))
        run_log('This program is only validated on DateTime version 4.3, You are using ' + str(pkg_resources.get_distribution("DateTime").version))
        print('')
        input("Press Enter to close the program window")
        os._exit(0)

    print('')
    print('System validated, Running program')
    print('')
    print(time.asctime())
    print('')
    
    run_log('System validated, Running program\n\n' + time.asctime() + '\n\n\n\n#######################################################################################\n\n')




#Find and Sort .TXT Raw Data Files
        
    excel_file_exists = False
    backup_file_exists = False
    Need_modified_quant_folder = True
    Need_modified_7500_folder = True
    Need_Project_Output_folder = True
    Plate_Info_Exists = False
    Re_Run_Sample_Info_Exists = False
    
    file_path_list = []


    #Data Directory Folder
    for root,dirs,files in os.walk(Data_Directory_Path):

        for names in files:
                
            if names.endswith('.txt') or names.endswith('.csv'):
                file_path_list.append(os.path.join(root,names))
                file_path_list.sort(key=os.path.getmtime)



    for root, dirs, files in os.walk(Project_Output_Path):

        if 'Modified_Quant_data_files' in dirs:
            Need_modified_quant_folder = False

        if 'Modified_7500_data_files' in dirs:
            Need_modified_7500_folder = False


        for names in files:

            if names == Excel_File_Name:
                excel_file_exists = True
            
            if names.startswith('Form10_BackUp'):
                Backup_Filename = os.path.join(root,names)
                backup_file_exists = True

            if names.split('\\')[-1] == 'Plate_Information.txt':
                Plate_Info_Exists = True
                
            if names.split('\\')[-1] == 'Re_Run_Sample_Info.txt':
                Re_Run_Sample_Info_Exists = True
                

    if Need_modified_quant_folder == True:
        os.mkdir(Project_Output_Path + '\\Modified_Quant_data_files')

    if Need_modified_7500_folder == True:
        os.mkdir(Project_Output_Path + '\\Modified_7500_data_files')


    
    file_count = 0





#Set Up Excel
        
    workbook_path = Project_Output_Path + '\\' + Excel_File_Name

    if excel_file_exists == True:
        wb = load_workbook(workbook_path, keep_vba=True)

    else:
        #SCS TEMPLATE LOCATION (Not DATA_DIRECTORY_PATH)
        wb = load_workbook(Excel_Template_Path, keep_vba=True)


    if 'Form10_Raw Data' in wb.sheetnames:
        ws = wb['Form10_Raw Data']
    else:
        ws = wb.create_sheet('Form10_Raw Data')

    if 'Form12_Plate Report' in wb.sheetnames:
        del wb['Form12_Plate Report']

    plate_report_ws = wb.create_sheet('Form12_Plate Report')

    if 'Form13_Re Run Samples' in wb.sheetnames:
        re_run_samples_ws = wb['Form13_Re Run Samples']
    else:
        re_run_samples_ws = wb.create_sheet('Form13_Re Run Samples')

    if 'Form14_AC Report' in wb.sheetnames:
        ac_report_ws = wb['Form14_AC Report']
    else:
        ac_report_ws = wb.create_sheet('Form14_AC Report')

    if 'Form15_Y Intercept Report' in wb.sheetnames:
        y_ws = wb['Form15_Y Intercept Report']
    else:
        y_ws = wb.create_sheet('Form15_Y Intercept Report')

    if 'Sheet1' in wb.sheetnames:
        del wb['Sheet1']
        


        
    #Exclude Plates
    exclude_plates_exists = False


    if os.path.exists(Project_Output_Path + '\\' + Excel_File_Name):    #VBA made the resulting workbook, excluded plate info is transferred
        exclude_plates = pd.read_excel(Project_Output_Path + '\\' + Excel_File_Name, sheet_name = 'Form2_Assay Information', header=0, usecols = ['Exclude these plates', 'Reason'])
    else:                                                               #VBA did not make the workbook, excluded plate info is within SCS_Path_that_ran_program ; whether that be the master template or a copy of it
        exclude_plates = pd.read_excel(SCS_Path_that_ran_program, sheet_name = 'Form2_Assay Information', header=0, usecols = ['Exclude these plates', 'Reason'])

    
    exclude_plates.dropna(inplace=True)
    exclude_plates.drop_duplicates(subset=['Exclude these plates'], inplace = True, ignore_index=True)
    if len(exclude_plates) > 0:
        exclude_plates_exists = True



    
#Set Up Informational .txt Files


    
    #Plate_Information.txt
    
    if Plate_Info_Exists == False: 
        with open(Project_Output_Path + '\\' + 'Plate_Information.txt', 'w', newline='') as Plate_Information:
            Plate_Information.write('Plate\tDate_Processed\tFile_Length\tPlate_Failed\tSamples_Failed\tAC_qty1\tAC_qty2\tY_Intercept\tSTD1_Ct_avg\tSTD4_Ct_avg\tSTD8_Ct_avg\tRun_Success')
        Plate_Information.close()
        
        Plate_Info_DF = pd.read_csv(Project_Output_Path + '\\' + 'Plate_Information.txt', sep = '\t', lineterminator = '\r', header = 0)

    else:
        Plate_Info_DF = pd.read_csv(Project_Output_Path + '\\' + 'Plate_Information.txt', sep = '\t', header = 0)

    
    files_before = len([i for i in Plate_Info_DF['Date_Processed'] if i.startswith('Excluded') == False])


    

    #Re_Run_Sample_Info
    
    if Re_Run_Sample_Info_Exists == False:
        with open(Project_Output_Path + '\\' + 'Re_Run_Sample_Info.txt', 'w') as Re_Run_Sample_Info:
            Re_Run_Sample_Info.write('Need_Re_Run_Plate\tNeed_Re_Run_Samples\tReason_Re_Run\tRe_Run_Plate\tRe_Run_Samples\tRun_Success')
        Re_Run_Sample_Info.close()

        Re_Run_Sample_Info_Dataframe = pd.read_csv(Project_Output_Path + '\\' + 'Re_Run_Sample_Info.txt', sep = '\t', lineterminator = '\n', header = 0)

    else:
        Re_Run_Sample_Info_Dataframe = pd.read_csv(Project_Output_Path + '\\' + 'Re_Run_Sample_Info.txt', sep = '\t', header = 0)





    #In case a file is removed from the 'Exclude Plates' list: Update the txt files

    previously_excluded_plates = []
    
    for i in range(len(Plate_Info_DF)):
        if Plate_Info_DF['Date_Processed'][i].startswith('Excluded'):
            previously_excluded_plates.append(Plate_Info_DF['Plate'][i])

    
    plate_info_drop_idx = []
    re_run_sample_drop_idx = []
    
    if len(previously_excluded_plates) > 0:
        for x in previously_excluded_plates:
            if x not in set(exclude_plates['Exclude these plates']):

                for z in range(len(Plate_Info_DF)):                    
                    if Plate_Info_DF['Plate'][z] == x:
                        plate_info_drop_idx.append(z)

                for y in range(len(Re_Run_Sample_Info_Dataframe)):
                    if Re_Run_Sample_Info_Dataframe['Need_Re_Run_Plate'][y] == x:
                        re_run_sample_drop_idx.append(y)


    Plate_Info_DF.drop(plate_info_drop_idx, inplace=True)
    Re_Run_Sample_Info_Dataframe.drop(re_run_sample_drop_idx, inplace=True)

    Plate_Info_DF.reset_index(inplace=True, drop=True)
    Re_Run_Sample_Info_Dataframe.reset_index(inplace=True, drop=True)




#Filter the Files that will get processed
    
    file_path_list = [i for i in file_path_list if i.split('\\')[-1][:-4] not in list(Plate_Info_DF['Plate'])]
    file_path_list = [i for i in file_path_list if i.split('\\')[-1] != 'Plate_Information.txt']
    file_path_list = [i for i in file_path_list if i.split('\\')[-1] != 'Re_Run_Sample_Info.txt']
    file_path_list = [i for i in file_path_list if i.split('\\')[-1].startswith('MODIFIED_QUANTSTUDIO_FILE') == False]
    file_path_list = [i for i in file_path_list if i.split('\\')[-1].startswith('MODIFIED_7500_FILE') == False]
    file_path_list = [i for i in file_path_list if i.split('\\')[-1].startswith('RE_ENCODED_7500_FILE') == False]
    file_path_list = [i for i in file_path_list if i.split('\\')[-1].endswith('log.txt') == False]
    file_path_list = [i for i in file_path_list if i.split('\\')[-1].startswith('Form1_Imported_Samples') == False]
    file_path_list = [i for i in file_path_list if i.split('\\')[-1].endswith('_Form4_Imported_ndv_list.txt') == False]
    file_path_list = [i for i in file_path_list if i.split('\\')[-1][:-4] not in set(exclude_plates['Exclude these plates'])]




#Exclude Plates Statement
    
    if exclude_plates_exists == True:
        
        print('')
        print('#######################################################################################')
        print('')
        print('Excluding Plates: ' + ', '.join([i for i in exclude_plates['Exclude these plates']]))
        run_log('Excluding Plates: ' + ', '.join([i for i in exclude_plates['Exclude these plates']]) + '\n\n')
        print('')
        
        for i in range(len(exclude_plates)):
            print('Excluding ' + exclude_plates['Exclude these plates'][i] + ' because: ' + exclude_plates['Reason'][i])
            run_log('Excluding ' + exclude_plates['Exclude these plates'][i] + ' because: ' + exclude_plates['Reason'][i] + '\n')
            
        run_log('\n')
        print('')
        print('#######################################################################################')
        print('')
        print('')
        print('Analyzing ' + str(len(file_path_list)) + ' Plates..')
        print('')
        run_log('#######################################################################################\n\n\n')
        run_log('Analyzing ' + str(len(file_path_list)) + ' Plates..\n\n\n')






        #Update Files -- new reasons
        for i in range(len(exclude_plates)):
            if exclude_plates['Exclude these plates'][i] in set(Plate_Info_DF['Plate']):

                for z in range(len(Plate_Info_DF)):
                    if Plate_Info_DF['Plate'][z] == exclude_plates['Exclude these plates'][i]:
                        Plate_Info_DF.iloc[z] = [exclude_plates['Exclude these plates'][i], 'Excluded because: ' + exclude_plates['Reason'][i],np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,True]
                        
                for q in range(len(Re_Run_Sample_Info_Dataframe)):
                    if Re_Run_Sample_Info_Dataframe['Need_Re_Run_Plate'][q] == exclude_plates['Exclude these plates'][i]:
                        Re_Run_Sample_Info_Dataframe.iloc[q] = [exclude_plates['Exclude these plates'][i],'All','Excluded because: ' + exclude_plates['Reason'][i],np.nan,np.nan,True]




        #Exclude Plates for first time
            else:
                Plate_Info_DF = Plate_Info_DF.append({'Plate': exclude_plates['Exclude these plates'][i], 'Date_Processed': 'Excluded because: ' + exclude_plates['Reason'][i], 'Run_Success':True}, ignore_index=True)
                Re_Run_Sample_Info_Dataframe = Re_Run_Sample_Info_Dataframe.append({'Need_Re_Run_Plate':exclude_plates['Exclude these plates'][i],'Need_Re_Run_Samples':'All','Reason_Re_Run': 'Excluded because: ' + exclude_plates['Reason'][i], 'Run_Success':True}, ignore_index=True)
                


    

#Early removal of failed plates from Plate_Info and Re_Run_Samples files to deal with unexpected program termination ex. Computer shutdown
    
    
    for i in range(len(Re_Run_Sample_Info_Dataframe)):                  #Mystery: Upon Reprocessing Re-Run-Sample-Info file, the Run_Success column is filled with 1.0 instead of True
        if Re_Run_Sample_Info_Dataframe.iloc[i,5] == 1.0:
            Re_Run_Sample_Info_Dataframe.iloc[i,5] = True
            
    Plate_Info_DF = Plate_Info_DF[Plate_Info_DF.Run_Success == True]
    Re_Run_Sample_Info_Dataframe = Re_Run_Sample_Info_Dataframe[Re_Run_Sample_Info_Dataframe.Run_Success == True]


    Plate_Info_DF.to_csv(Project_Output_Path + '\\' + 'Plate_Information.txt', sep='\t', index = False, header = True, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')




        
#Initialize variables outside the data file loop
    
    row_where_file_ends = len([i for i in Plate_Info_DF['Date_Processed'] if i.startswith('Excluded') == False]) * 109           #Length of Plate_Info_DF that isn't excluded plates
    currentRow = 1 + row_where_file_ends
    
    plates_passed = 0 #For AC report

    AC_qty_means = []

    for plate in Plate_Info_DF['Plate_Failed']:
        if plate == False:
            plates_passed += 1





#Make sure files are closed before writing to them
        
    if excel_file_exists == True:
        while True: 
            try:
                with open(workbook_path, 'r+'):            
                    break
            except IOError:
                input("The program cannot save to the SCS19 file while it is open. Please close it and Press Enter to retry")

    if backup_file_exists == True:
        while True: 
            try:
                with open(Backup_Filename, 'r+'):            
                    break
            except IOError:
                input("The program cannot save to the Back up SCS19 file while it is open. Please close it and Press Enter to retry")


   

#Eliminate the 1 row off issue and Welcome the next batch
    
    if row_where_file_ends > 0:
        ws['A' + str(row_where_file_ends)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        ws['B' + str(row_where_file_ends)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        ws['C' + str(row_where_file_ends)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        ws['D' + str(row_where_file_ends)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        ws['E' + str(row_where_file_ends)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        ws['F' + str(row_where_file_ends)] = 'Files Processed ' + str(now.strftime('%Y-%m-%d %H:%M')) + ' and after'



#Iterate over every file


    for data_file in file_path_list:



        if cause_error == True:
            if data_file == file_path_list[5]:
                #os._exit(0)
                print('Hey Xuening can I have database access?' + 1)



              
#Classify the Data file

        is_7500 = False
        
        try:
            test_the_file = pd.read_csv(data_file, sep = '\t', skip_blank_lines = False, usecols = [0])
        except:
            is_7500 = True


        if is_7500 == False:
            
            if test_the_file.columns[0] == 'SDS 2.4': #Hello 7900


                print('')
                print('Processing 7900 File: ' + data_file.split('\\')[-1])
                run_log(time.asctime() + '\n')
                run_log('Processing 7900 File: ' + data_file.split('\\')[-1] + '\n')

                form10 = pd.read_csv(data_file, sep = '\t', lineterminator='\n', header = 10)
                form10_raw_extra = pd.read_csv(data_file, sep = '\t', lineterminator='\n', nrows = 9)

                    
#Clean Raw Data
            
                form10.drop(['\r','FOS','HMD','LME','EW','BPR','NAW','HNS','HRN','EAF','BAF','TAF','CAF'], axis = 1, inplace = True)


                
#Write Raw Data to Excel

                open_data_file = open(data_file, 'r')

                currentRow = (file_count + files_before) * 109 + 1
                currentCell = None
                currentCol = 2

                while True:
                
                    line = open_data_file.readline()

                    if not line:
                        break

                    if line in ['\n', '\r', '\r\n'] or line.startswith('NAP'):
                        continue

                    fields = line.split("\t") #separate line by tabs into up to 34 cells
                    for fieldIndex in range(len(fields)): #copy up to 21 cells in each line
                    
                        currentCell = ws.cell(row=currentRow, column=currentCol)

                        if fields[fieldIndex]: #only try to copy field if it has a value
                            if (fields[fieldIndex][0].isdigit() or fields[fieldIndex].startswith("-")): #set cell value as a float if it begins with a number or a "-"

                                try: #copy value as a float
                                    currentCell.value = float(fields[fieldIndex])

                                except: #copy value as text
                                    currentCell.value = fields[fieldIndex]

                            else: #copy value as text
                                currentCell.value = fields[fieldIndex]

                        currentCol += 1

                        if (currentCol == 23 or fieldIndex == len(fields)-1): #move to the next row in sheet after 21 columns or at the end of the row
                            currentCol = 2
                            currentRow += 1
                            break

                open_data_file.close()

                print('7900 file: ' + data_file.split('\\')[-1] + ' is written')
                run_log('7900 file: ' + data_file.split('\\')[-1] + ' is written\n')


#More Cleaning: Drop the bottom (floats), set the slope, R2, Drop nan samples
                

                form10['Slope'] = float(form10.iloc[-8,1])
                form10['R^2'] = float(form10.iloc[-6,2][:-1])               #[:-1] gets rid of '/r' in the value
                form10['Y_Intercept'] = float(form10.iloc[-7,2][:-1])
                
                nan_indexes = []

                for x in range(0,len(form10)):
                    if form10['Well'][x] == 'Slope':
                        break

                form10.drop(form10.tail(len(form10)-x).index, inplace = True)


                for x in range(0,len(form10)):
                    if pd.isna(form10['Sample Name'][x]) == True:
                        nan_indexes.append(x)


                form10.drop(index = nan_indexes, inplace = True)
                form10.reset_index(inplace = True)


                
                #Rename SPK, Std, AC
                spike_end = 0
                Titer_info_exists = False
                


                for i in range(len(form10)):
                        
                    if form10['Sample Name'][i].upper() == 'SPK CONTROL':
                        form10.iloc[i, 2] = 'SPK CONTROL'
                    
                    if form10['Sample Name'][i].lower() == 'spike control':
                        form10.iloc[i, 2] = 'SPK CONTROL'

                    if form10['Sample Name'][i] != form10['Sample Name'][i].upper():
                        form10.iloc[i, 2] = form10.iloc[i, 2].upper()

                    if form10['Sample Name'][i].lower() == 'assay control' or form10['Sample Name'][i].lower() == 'assaycontrol':
                        form10.iloc[i, 2] = 'AC'

                    if form10['Sample Name'][i] != form10['Sample Name'][i].replace(' ', '') and form10['Sample Name'][i].startswith('SP') == False:
                        form10.iloc[i, 2] = form10['Sample Name'][i].replace(' ', '')

                    if form10['Sample Name'][i].startswith('SP') == False:
                        form10.iloc[i, 2] = str(form10['Sample Name'][i]).strip()

                form10 = form10.sort_values(by=['Sample Name'], ignore_index = True)





################################################################################################################################################################



            elif '*' in test_the_file.columns[0]:   # Hello Quant Studio

                print('')
                print('Processing Quant Studio File: ' + data_file.split('\\')[-1])
                run_log(time.asctime() + '\n')
                run_log('Processing Quant Studio File: ' + data_file.split('\\')[-1] +'\n')

                temp_form10 = pd.read_csv(data_file, sep = '\t', skip_blank_lines = False, usecols = [0])
                Index_where_data_begins = 0
                Data_exists = False
                is_QS = True


#Find the actual data
                
                for i in range(0, len(temp_form10)):
                    if temp_form10.iloc[:,0][i] == '[Results]':
                        Data_exists = True
                        Index_where_data_begins = i+2 #+2 because Row 0 is the header and Row 1 is index 0

                if Data_exists == False:
                    
                    print("Theres no Data in QS File" + data_file.split('\\')[-1] + ", Plate " + str(file_count+files_before+1) + " will be empty")
                    run_log("Theres no Data in QS File" + data_file.split('\\')[-1] + ", Plate " + str(file_count+files_before+1) + " will be empty")
                    print('')
                    
                    if row_where_file_ends == 0:
                        ws['U' + str(row_where_file_ends+1)] = str(1 + file_count + files_before)
                        ws['U' + str(row_where_file_ends+1)].border = Border(bottom = Side(style = 'medium'))
                        ws['U' + str(row_where_file_ends+1)].font = Font(size = 16, bold = True)
                        ws['U' + str(row_where_file_ends+1)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
                        ws['U' + str(row_where_file_ends+1)].alignment = Alignment(horizontal = 'center')
                        ws['V' + str(row_where_file_ends+1)] = 'Empty Quant Studio File'
                        ws['V' + str(row_where_file_ends+1)].font = Font(bold = True)

                        row_where_file_ends += 1 #Avoid error of pasting into row 0

                        with open(Data_Directory_Path + '\\' + 'Plate_Information.txt', 'a', newline = '\r') as Plate_Information:
                            Plate_Information.write('\n'+data_file.split('\\')[-1][:-4]+'\t'+str(now.strftime('%Y-%m-%d %H:%M'))+'\t'+str(1)+'\t\t\t\t\tFalse')
                        Plate_Information.close()



                        #Plate Report

                        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
                    
                        plate_report_ws['A'+ str(Paste_row)] = 'Plate ' + str(file_count+1)
                        plate_report_ws['A' + str(Paste_row+1)] = data_file.split('\\')[-1][:-4]
                        plate_report_ws['A'+ str(Paste_row+1)].font = Font(bold = True)
                        plate_report_ws['A'+ str(Paste_row)].font = Font(bold = True)
                        plate_report_ws['B'+ str(Paste_row)] = 'Re-Run Status'
                        plate_report_ws['B'+ str(Paste_row+1)].fill = PatternFill(start_color='00FF0000',end_color='00FF0000',fill_type='solid')
                        plate_report_ws['D'+ str(Paste_row)] = 'Reason'
                        plate_report_ws['D'+ str(Paste_row+1)] = 'Empty Quant Studio File'

                    else:
                        ws['U' + str(row_where_file_ends)] = str(1 + file_count + files_before)
                        ws['U' + str(row_where_file_ends)].border = Border(bottom = Side(style = 'medium'))
                        ws['U' + str(row_where_file_ends)].font = Font(size = 16, bold = True)
                        ws['U' + str(row_where_file_ends)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
                        ws['U' + str(row_where_file_ends)].alignment = Alignment(horizontal = 'center')
                        ws['V' + str(row_where_file_ends)] = 'Empty Quant Studio File'

                        #Plate Report

                        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
                    
                        plate_report_ws['A'+ str(Paste_row)] = 'Plate ' + str(file_count+1)
                        plate_report_ws['A' + str(Paste_row+1)] = data_file.split('\\')[-1][:-4]
                        plate_report_ws['A'+ str(Paste_row+1)].font = Font(bold = True)
                        plate_report_ws['A'+ str(Paste_row)].font = Font(bold = True)
                        plate_report_ws['B'+ str(Paste_row)] = 'Re-Run Status'
                        plate_report_ws['B'+ str(Paste_row+1)].fill = PatternFill(start_color='00FF0000',end_color='00FF0000',fill_type='solid')
                        plate_report_ws['D'+ str(Paste_row)] = 'Reason'
                        plate_report_ws['D'+ str(Paste_row+1)] = 'Empty Quant Studio File'



#Process Quant File
                    
                else:
                    
                    form10_unsorted = pd.read_csv(data_file, sep = '\t', lineterminator = '\n', header = Index_where_data_begins)#, usecols = ['Well', 'Sample Name', 'Target Name','Reporter','Task','CT','Quantity','Quantity Mean', 'Slope','R(superscript 2)', 'Y-Intercept'])

                    if 'Y-Intercept' in list(form10_unsorted.columns):
                        form10_unsorted = pd.read_csv(data_file, sep = '\t', lineterminator = '\n', header = Index_where_data_begins, usecols = ['Well', 'Sample Name', 'Target Name','Reporter','Task','CT','Quantity','Quantity Mean', 'Slope','R(superscript 2)', 'Y-Intercept'])

                    else:
                        form10_unsorted = pd.read_csv(data_file, sep = '\t', lineterminator = '\n', header = Index_where_data_begins, usecols = ['Well', 'Sample Name', 'Target Name','Reporter','Task','CT','Quantity','Quantity Mean', 'Slope','R(superscript 2)'])



                    #Rename SPK, Std, AC
                    spike_end = 0
                    Titer_info_exists = False



                    
                    for i in range(len(form10_unsorted)):

                        if pd.isna(form10_unsorted['Sample Name'][i]) == True:
                            continue
                            
                        if form10_unsorted['Sample Name'][i].upper() == 'SPK CONTROL':
                            form10_unsorted.iloc[i, 1] = 'SPK CONTROL'    #1 has the sample names
                        
                        if form10_unsorted['Sample Name'][i].lower() == 'spike control':
                            form10_unsorted.iloc[i, 1] = 'SPK CONTROL'

                        if form10_unsorted['Sample Name'][i] != form10_unsorted['Sample Name'][i].upper():
                            form10_unsorted.iloc[i, 1] = form10_unsorted.iloc[i, 1].upper()

                        if form10_unsorted['Sample Name'][i] != form10_unsorted['Sample Name'][i].replace(' ', '') and form10_unsorted['Sample Name'][i].startswith('SP') == False:
                            form10_unsorted.iloc[i, 1] = form10_unsorted['Sample Name'][i].replace(' ', '')

                        if form10_unsorted['Sample Name'][i].lower() == 'assay control' or form10_unsorted['Sample Name'][i].lower() == 'assaycontrol':
                            form10_unsorted.iloc[i, 1] = 'AC'

                        if form10_unsorted['Sample Name'][i].startswith('SP') == False:
                            form10_unsorted.iloc[i, 1] = str(form10_unsorted['Sample Name'][i]).strip()

                        if form10_unsorted['Sample Name'][i].upper().startswith('TITER') == True:
                            Titer_info_exists = True

            

                    #Sort Samples
                    form10_full = form10_unsorted.sort_values(by=['Sample Name'], ignore_index = True)

                            
                    #Rename Columns
                    form10_full.rename(columns={"CT": "Ct", "Quantity Mean": "Qty Mean", "R(superscript 2)":"R^2"}, inplace = True)

                    if 'Y-Intercept' in list(form10_full.columns):
                        form10_full.rename(columns={"Y-Intercept": "Y_Intercept"}, inplace = True)

                    #Take commas out of columns
                    form10_full['Quantity'] = form10_full['Quantity'].str.replace(',', '').astype('float')
                    form10_full['Qty Mean'] = form10_full['Qty Mean'].str.replace(',', '').astype('float')

                        
                    #Chop off Titer Information
                    
                    if Titer_info_exists == True:
                        
                        titer_idx = 0
                        titers = []
                        for i in range(len(form10_full['Sample Name'])):
                            if form10_full['Sample Name'][i].upper().startswith('TITER'):
                                titer_idx = i
                                break

                        for i in range(titer_idx, len(form10_full)):
                            titers.append(i)

                        form10 = form10_full.drop(titers)

                    else:
                        form10 = form10_full
                        

                    #Add a Header/Footer
                    Header = pd.DataFrame()
                    Header = Header.append(['Quant Studio Results',' ','For File:',data_file.split('\\')[-1][:-4],' ',' ',' ',' ', 'Sample Information',' '])
                    Footer = pd.DataFrame()
                    Footer = Footer.append([' ',' '])


#Create Modified Quant Studio File

                    Header.to_csv(Project_Output_Path + '\\Modified_Quant_data_files\\MODIFIED_QUANTSTUDIO_FILE_' + data_file.split('\\')[-1], sep = '\t', mode='w',index=False, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')
                    form10.to_csv(Project_Output_Path + '\\Modified_Quant_data_files\\MODIFIED_QUANTSTUDIO_FILE_' + data_file.split('\\')[-1], sep = '\t', mode='a',index=False, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')
                    Footer.to_csv(Project_Output_Path + '\\Modified_Quant_data_files\\MODIFIED_QUANTSTUDIO_FILE_' + data_file.split('\\')[-1], sep = '\t', mode='a',index=False, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')
                    

#Write QS Data to Excel

                    open_data_file = open(Project_Output_Path + '\\Modified_Quant_data_files\\MODIFIED_QUANTSTUDIO_FILE_' + data_file.split('\\')[-1], 'r')
                    
                    currentRow = (file_count + files_before) * 109 + 1
                    currentCell = None
               
                    currentCol = 2

                    while True:
                        
                        line = open_data_file.readline()

                        if not line:
                            break

                        if line in ['\n', '\r', '\r\n'] or line.startswith('NAP'):
                            continue

                        fields = line.split("\t") #separate line by tabs into up to 34 cells
                        for fieldIndex in range(len(fields)): #copy up to 21 cells in each line
                        
                            currentCell = ws.cell(row=currentRow, column=currentCol)

                            if fields[fieldIndex]: #only try to copy field if it has a value
                                if (fields[fieldIndex][0].isdigit() or fields[fieldIndex].startswith("-")): #set cell value as a float if it begins with a number or a "-"

                                    try: #copy value as a float
                                        currentCell.value = float(fields[fieldIndex])

                                    except: #copy value as text
                                        currentCell.value = fields[fieldIndex]

                                else: #copy value as text
                                    currentCell.value = fields[fieldIndex]

                            currentCol += 1

                            if (currentCol == 23 or fieldIndex == len(fields)-1): #move to the next row in sheet after 21 columns or at the end of the row
                                currentCol = 2
                                currentRow += 1
                                break


                    open_data_file.close()
                        

                    print('Quant Studio file ' + data_file.split('\\')[-1] + ' is written')
                    run_log('Quant Studio file ' + data_file.split('\\')[-1] + ' is written\n')
                    


#More Cleaning: Drop nan samples
                     
                    nan_indexes = []
                    
                        
                    for x in range(0,len(form10)):
                        if pd.isna(form10['Sample Name'][x]) == True:
                            nan_indexes.append(x)

                    form10.drop(index = nan_indexes, inplace = True)
                    form10.reset_index(inplace = True)


##########################################################################################################################################################################
                
        if is_7500 == True: #Hello 7500
            
            
            print('')
            print('Processing 7500 File: ' + data_file.split('\\')[-1])
            run_log(time.asctime() + '\n')
            run_log('Processing 7500 File: ' + data_file.split('\\')[-1] + '\n')


            with open(data_file, 'r', encoding='utf-8', errors='ignore') as infile, open(Project_Output_Path + '\\Modified_7500_data_files\\RE_ENCODED_7500_FILE_' + data_file.split('\\')[-1], 'w') as outfile:
                inputs = csv.reader(infile)
                output = csv.writer(outfile)
    
                for index, row in enumerate(inputs):
                    output.writerow(row)

            infile.close()
            outfile.close()
            

            find_form10_data = pd.read_csv(Project_Output_Path + '\\Modified_7500_data_files\\RE_ENCODED_7500_FILE_' + data_file.split('\\')[-1], usecols = [0])
            Index_where_data_begins = 0
            Index_where_slope_r_y_is = 0
            

            #Find the actual data
            
            for i in range(0, len(find_form10_data)):
                if find_form10_data.iloc[i,0] == 'Well':
                    Index_where_data_begins = i+1
                if find_form10_data.iloc[i,0] == 'Detector Name':
                    Index_where_slope_r_y_is = i+1



            #Read File
            form10_slope_stuff = pd.read_csv(Project_Output_Path + '\\Modified_7500_data_files\\RE_ENCODED_7500_FILE_' + data_file.split('\\')[-1], header = Index_where_slope_r_y_is)#, nrows = 2)    #, usecols = ['Slope', 'Intercept', 'R2'])
            form10_unsorted = pd.read_csv(Project_Output_Path + '\\Modified_7500_data_files\\RE_ENCODED_7500_FILE_' + data_file.split('\\')[-1], header = Index_where_data_begins, usecols = ['Well', 'Sample Name', 'Detector', 'Task','Ct','Qty','Mean Qty'])


            
            #Rename SPK, Std, AC
            
            for i in range(len(form10_unsorted)):

                if pd.isna(form10_unsorted['Sample Name'][i]) == True:
                    continue
                    
                if form10_unsorted['Sample Name'][i].upper() == 'SPK CONTROL':
                    form10_unsorted.iloc[i, 1] = 'SPK CONTROL'    #1 has the sample names
                
                if form10_unsorted['Sample Name'][i].lower() == 'spike control':
                    form10_unsorted.iloc[i, 1] = 'SPK CONTROL'

                if form10_unsorted['Sample Name'][i] != form10_unsorted['Sample Name'][i].upper():
                    form10_unsorted.iloc[i, 1] = form10_unsorted.iloc[i, 1].upper()

                if form10_unsorted['Sample Name'][i] != form10_unsorted['Sample Name'][i].replace(' ', '') and form10_unsorted['Sample Name'][i].startswith('SP') == False:
                    form10_unsorted.iloc[i, 1] = form10_unsorted['Sample Name'][i].replace(' ', '')

                if form10_unsorted['Sample Name'][i].lower() == 'assay control' or form10_unsorted['Sample Name'][i].lower() == 'assaycontrol':
                    form10_unsorted.iloc[i, 1] = 'AC'

                if form10_unsorted['Sample Name'][i].startswith('SP') == False:
                    form10_unsorted.iloc[i, 1] = str(form10_unsorted['Sample Name'][i]).strip()


        

            #Sort Samples
            form10=form10_unsorted.sort_values(by=['Sample Name'], ignore_index = True)

                    
            #Manage Columns
            form10.rename(columns={"Qty": "Quantity", "Mean Qty": "Qty Mean"}, inplace = True)

            #form10.drop(['StdDev Qty', 'Filtered', 'Tm', 'User Defined #1', 'User Defined #2', 'User Defined #3'], axis = 1, inplace = True)

            form10['Slope'] = pd.Series(form10_slope_stuff['Slope'][0])
            form10['Y_Intercept'] = pd.Series(form10_slope_stuff['Intercept'][0])
            form10['R^2'] = pd.Series(form10_slope_stuff['R2'][0])

                

            #Add a Header/Footer
            Header = pd.DataFrame()
            Header = Header.append(['7500 Results',' ','For File:',data_file.split('\\')[-1][:-4],' ',' ',' ',' ', 'Sample Information',' '])
            Footer = pd.DataFrame()
            Footer = Footer.append([' ',' '])
            

#Create Modified 7500 File

            Header.to_csv(Project_Output_Path + '\\Modified_7500_data_files\\MODIFIED_7500_FILE_' + data_file.split('\\')[-1], sep = '\t', mode='w',index=False, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')
            form10.to_csv(Project_Output_Path + '\\Modified_7500_data_files\\MODIFIED_7500_FILE_' + data_file.split('\\')[-1], sep = '\t', mode='a',index=False, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')
            Footer.to_csv(Project_Output_Path + '\\Modified_7500_data_files\\MODIFIED_7500_FILE_' + data_file.split('\\')[-1], sep = '\t', mode='a',index=False, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')
            

#Write 7500 Data to Excel

            open_data_file = open(Project_Output_Path + '\\Modified_7500_data_files\\MODIFIED_7500_FILE_' + data_file.split('\\')[-1], 'r')
            
            currentRow = (file_count + files_before) * 109 + 1
            currentCell = None
       
            currentCol = 2

            while True:
                
                line = open_data_file.readline()

                if not line:
                    break

                if line in ['\n', '\r', '\r\n'] or line.startswith('NAP'):
                    continue

                fields = line.split("\t") #separate line by tabs into up to 34 cells
                for fieldIndex in range(len(fields)): #copy up to 21 cells in each line
                
                    currentCell = ws.cell(row=currentRow, column=currentCol)

                    if fields[fieldIndex]: #only try to copy field if it has a value
                        if (fields[fieldIndex][0].isdigit() or fields[fieldIndex].startswith("-")): #set cell value as a float if it begins with a number or a "-"

                            try: #copy value as a float
                                currentCell.value = float(fields[fieldIndex])

                            except: #copy value as text
                                currentCell.value = fields[fieldIndex]

                        else: #copy value as text
                            currentCell.value = fields[fieldIndex]

                    currentCol += 1

                    if (currentCol == 23 or fieldIndex == len(fields)-1): #move to the next row in sheet after 21 columns or at the end of the row
                        currentCol = 2
                        currentRow += 1
                        break


            open_data_file.close()
                

            print('7500 file ' + data_file.split('\\')[-1] + ' is written')
            run_log('7500 file ' + data_file.split('\\')[-1] + ' is written\n')
            

#More Cleaning: Drop nan samples
             
            nan_indexes = []

            for x in range(0,len(form10)):
                if pd.isna(form10['Sample Name'][x]) == True:
                    nan_indexes.append(x)

            form10.drop(index = nan_indexes, inplace = True)
            form10.reset_index(inplace = True)


            
##########################################################################################################################################################################
            
#Form 10 Table
            

        #Manage Paste Offset

        Paste_row = row_where_file_ends + 20

        Sample_Name_List = Get_Sample_Names(form10)

        AC_exist = False
        for name in Sample_Name_List:
            if name.startswith('AC'):
                Paste_row -= 1
                AC_exist = True
                break
        

        #Paste Calculation Header
        
        ws['X' + str(Paste_row-1)] = 'Sample Name'
        
        ws['Y' + str(Paste_row-1)] = 'Qty Mean'
        
        ws['Z' + str(Paste_row-1)] = 'Replicate Difference'
        
        ws['AA' + str(Paste_row-1)] = '(+) SPK Value'
        ws['AB' + str(Paste_row-1)] = 'SPK Difference'
        ws['AC' + str(Paste_row-1)] = 'Inhibition?'
        
        #Paste Sample Names
        
        for name in Sample_Name_List:
            ws['X' + str(Paste_row)] = name
            Paste_row += 1
        

        #QuantityMean Calculation and write
            
        Paste_row = row_where_file_ends + 20 #Reset Paste_Row position

        if AC_exist == True:
            Paste_row -= 1
         

        QuantityMeans = QtyMean_Dataframe(form10)
        
        for i in range(len(QuantityMeans['Qty Mean'])):
            
            if QuantityMeans['Sample Name'][i] == 'STD1': #STD curve range
                Paste_row += 3
                
            ws['Y' + str(Paste_row)] = QuantityMeans['Qty Mean'][i]
            Paste_row += 1
    

        #Replicate Diff
        Paste_row = row_where_file_ends + 20 #Reset Paste_Row position

        if AC_exist == True:
            Paste_row -= 1
            
        Replicate_Differences = Replicate_Diff_Dataframe(form10, LOQ)
        for value in Replicate_Differences['Ct diff']:
            ws['Z' + str(Paste_row)] = value
            Paste_row+=1
            

        #SPK Ct
        Paste_row = row_where_file_ends + 20 #Reset Paste_Row position

        if AC_exist == True:
            Paste_row -= 1
            
        if Sample_Name_List[0].startswith('A'):
            Paste_row += 1


            
        SPK_Ct = SPK_Ct_dataframe(form10) #SPK_Ct will be used for a list of C_Samples
        for x in range(len(SPK_Ct['SPK Values'])):
            ws['AA' + str(Paste_row)] = SPK_Ct['SPK Values'][x]
            ws['AB' + str(Paste_row)] = SPK_Ct['SPK Value Differences'][x]
            ws['AC' + str(Paste_row)] = SPK_Ct['Inhibition?'][x]
            Paste_row += 1


        #SPK Control Information
        Paste_row += 2
        ws['AA' + str(Paste_row)] = SPK_Ct['SPK Control Value'][0]
        ws['AA' + str(Paste_row)].font = Font(size = 13, bold = True)


#Style Formatting
        
        #ws['U' + str(row_where_file_ends+1)] = 1 + file_count + files_before
        #ws['U' + str(row_where_file_ends+1)].font = Font(size = 16, bold = True)
        #ws['U' + str(row_where_file_ends+1)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
        #ws['U' + str(row_where_file_ends+1)].alignment = Alignment(horizontal = 'center')

        row_where_file_ends = row_where_file_ends + 109

###############################################################################################################################################################################

        
#Plate Report

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
        
        plate_report_ws['A'+ str(Paste_row)] = 'Plate ' + str(file_count+1)
        plate_report_ws['B'+ str(Paste_row)] = 'Re-Run Status'
        plate_report_ws['C'+ str(Paste_row)] = 'Samples'
        plate_report_ws['D'+ str(Paste_row)] = 'Reason'
        plate_report_ws['F'+ str(Paste_row)] = 'Table'

        plate_report_ws['A' + str(Paste_row+1)] = data_file.split('\\')[-1][:-4]
        plate_report_ws['A'+ str(Paste_row+1)].font = Font(bold = True)
        plate_report_ws['A' + str(Paste_row+1)].fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')

        plate_report_ws['A'+ str(Paste_row)].font = Font(bold = True)




#Plate Report Table
        
        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
        
        plate_report_ws['F' + str(Paste_row+1)] = 'Sample Name'
        
        plate_report_ws['G' + str(Paste_row+1)] = 'Qty Mean'
        
        plate_report_ws['H' + str(Paste_row+1)] = 'Replicate Difference'
        
        plate_report_ws['I' + str(Paste_row+1)] = '(+) SPK Value'
        plate_report_ws['J' + str(Paste_row+1)] = 'SPK Difference'
        plate_report_ws['K' + str(Paste_row+1)] = 'Inhibition?'
        

        for name in Sample_Name_List:
            plate_report_ws['F' + str(Paste_row+2)] = name
            Paste_row += 1

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position


        for i in range(len(QuantityMeans['Qty Mean'])):
            
            if QuantityMeans['Sample Name'][i] == 'STD1': #STD curve range
                Paste_row += 3

            plate_report_ws['G' + str(Paste_row+2)] = QuantityMeans['Qty Mean'][i]
            Paste_row += 1

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
            
        for value in Replicate_Differences['Ct diff']:
            plate_report_ws['H' + str(Paste_row+2)] = value
            Paste_row+=1

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position

        if Sample_Name_List[0].startswith('AC'):
            Paste_row += 1
        
        for x in range(len(SPK_Ct['SPK Values'])):
            plate_report_ws['I' + str(Paste_row+2)] = SPK_Ct['SPK Values'][x]
            plate_report_ws['J' + str(Paste_row+2)] = SPK_Ct['SPK Value Differences'][x]
            plate_report_ws['K' + str(Paste_row+2)] = SPK_Ct['Inhibition?'][x]
            Paste_row += 1


        Paste_row += 2
        plate_report_ws['I' + str(Paste_row+2)] = SPK_Ct['SPK Control Value'][0]
        plate_report_ws['I' + str(Paste_row+2)].font = Font(size = 12, bold = True)

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
        



#ReRun Status (Plates)

        re_run_status = 'green'
        re_run_samples_replicates = 'not yet'
        re_run_samples_inhib = 'not yet'
        
        Plate_Fail = False #Lets track if the plate failed
        Samples_Failed = False #and if samples failed
        

        #Status Color Indicator Default is Green
        plate_report_ws['B'+ str(Paste_row+1)].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

        

        #STD 1-7, STD 8, & Grabbing all samples needing re run due to STD

        plate_fail_reason = []
        sample_fail_reason = []
        STD_additional_comment = []
        
        
        for x in range(0,len(Replicate_Differences['Sample Name'])): #Different repl diff criteria for std 1-7 and std 8

            
            
            if Replicate_Differences['Sample Name'][x] != 'STD8' and Replicate_Differences['Sample Name'][x].startswith('STD'): #STD 1-7
                    
                if str(Replicate_Differences['Ct diff'][x]).startswith('PROBLEM') or '.' in str(Replicate_Differences['Ct diff'][x]) and abs(Replicate_Differences['Ct diff'][x]) > Repl_Diff_Criteria_for_STD_1_to_7: #Find Problems 

                        plate_fail_reason.append(Replicate_Differences['Sample Name'][x])


                        plate_report_ws['B'+ str(Paste_row+2)] = 'Re-Run Plate'
                        plate_report_ws['D'+ str(Paste_row+2)] = Replicate_Differences['Sample Name'][x] + ' Failed' + ''.join(STD_additional_comment)

                        if len(STD_additional_comment) == 0:
                            STD_additional_comment.append(Replicate_Differences['Sample Name'][x] + ' Failed') #STD 1-7 message
                        else:
                            STD_additional_comment.append(', ' + Replicate_Differences['Sample Name'][x] + ' Failed')
                            
                        re_run_status = 'red'
                        Samples_Failed = True   #apparently std 1-7 dont fail plate?
                        

                            
            if Replicate_Differences['Sample Name'][x] == 'STD8': #STD 8

                if str(Replicate_Differences['Ct diff'][x]).startswith('PROBLEM') or '.' in str(Replicate_Differences['Ct diff'][x]) and abs(Replicate_Differences['Ct diff'][x]) > Repl_Diff_Criteria_for_STD_8: #Find Problems
                    
                        plate_fail_reason.append(Replicate_Differences['Sample Name'][x])

                            
                        plate_report_ws['B'+ str(Paste_row+2)] = 'Re-Run Plate'
                        plate_report_ws['D'+ str(Paste_row+2)] = 'STD 8 Failed' + ''.join(STD_additional_comment)

                        re_run_status = 'red'
                        Plate_Fail = True
                        Samples_Failed = True
                            

                
        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
                                             



        #Slope, R2, LOD : The plate requirements

        plate_report_ws['M' + str(Paste_row+1)] = 'Slope'
        plate_report_ws['N' + str(Paste_row+1)] = 'R^2'
        plate_report_ws['O' + str(Paste_row+1)] = 'NTC'
        plate_report_ws['P' + str(Paste_row+1)] = 'LOD'
        plate_report_ws['Q' + str(Paste_row+1)] = 'LOQ'
        plate_report_ws['M' + str(Paste_row+2)] = str(form10['Slope'][0]) + ' cycles/log decade'
        plate_report_ws['N' + str(Paste_row+2)] = form10['R^2'][0]
            
        plate_report_ws['P' + str(Paste_row+2)] = LOD
        plate_report_ws['Q' + str(Paste_row+2)] = LOQ

        
        if QuantityMeans['NTC'][0] == 'Undetermined':
            plate_report_ws['O' + str(Paste_row+2)] = 'Undetermined'
        else:
            plate_report_ws['O' + str(Paste_row+2)] = QuantityMeans['NTC'][0]


        if float(form10['Slope'][0]) < float(SlopeMin):
            plate_report_ws['B'+ str(Paste_row+2)] = 'Re-Run Plate'
            plate_report_ws['D'+ str(Paste_row+3)] = 'Slope Requirement'
            plate_fail_reason.append('Slope Requirement')
            re_run_status = 'red'
            Plate_Fail = True
            Samples_Failed = True

        if float(form10['Slope'][0]) > float(SlopeMax):
            plate_report_ws['B'+ str(Paste_row+2)] = 'Re-Run Plate'
            plate_report_ws['D'+ str(Paste_row+3)] = 'Slope Requirement'
            plate_fail_reason.append('Slope Requirement')
            re_run_status = 'red'
            Plate_Fail = True
            Samples_Failed = True
            
        if isinstance(form10['R^2'][0], str) == False:
            if float(form10['R^2'][0]) < float(R2):
                plate_report_ws['B'+ str(Paste_row+2)] = 'Re-Run Plate'
                plate_report_ws['D'+ str(Paste_row+4)] = 'R^2 Requirement'
                plate_fail_reason.append('R2 Requirement')
                re_run_status = 'red'
                Plate_Fail = True 
         
        if QuantityMeans['NTC'][0] != 'Undetermined':
            if QuantityMeans['NTC'][0] < float(LOD) and (math.isnan(QuantityMeans['NTC'][0]) == False) and QuantityMeans['NTC'][0] != 0:
                plate_report_ws['B'+ str(Paste_row+2)] = 'Re-Run Plate'
                plate_report_ws['D'+ str(Paste_row+5)] = 'NTC vs LOD Requirement'
                plate_fail_reason.append('NTC vs LOD Requirement')
                re_run_status = 'red'
                Plate_Fail = True
                Samples_Failed = True


        
######################################################################################################################################################################

#Re_Run_Samples Report

        
        #Connect Samples that were re_run with their former selves.
        for C_name in SPK_Ct['Sample Name']:
            C_name_with_no_plus_SPK = C_name.split('+')[0]
            C_num = C_name.split('_')[0]
            for Re_Run_Sample_Info_Index in range(0,len(Re_Run_Sample_Info_Dataframe['Need_Re_Run_Samples'])):

                if Re_Run_Sample_Info_Dataframe['Need_Re_Run_Samples'][Re_Run_Sample_Info_Index] != np.nan or Re_Run_Sample_Info_Dataframe['Need_Re_Run_Samples'][Re_Run_Sample_Info_Index] != 'All':
                    
                    if C_num == Re_Run_Sample_Info_Dataframe['Need_Re_Run_Samples'][Re_Run_Sample_Info_Index].split('_')[0]:
                        

                        #Robust dumb method FOR PASTING RE RAN SAMPLES into DF, relies on date sort for accuracy
                        Re_Run_Sample_Info_Dataframe.iloc[Re_Run_Sample_Info_Index, 3] = data_file.split('\\')[-1][:-4]
                        Re_Run_Sample_Info_Dataframe.iloc[Re_Run_Sample_Info_Index, 4] = C_name_with_no_plus_SPK
                    


        #Plate Failiure check: AC Diff
        for Replicate_value_index in range(0,len(Replicate_Differences)):            

            if str(Replicate_Differences['Ct diff'][Replicate_value_index]).startswith('PROBLEM'):

                if Replicate_Differences['Sample Name'][Replicate_value_index].startswith('AC'):
                    plate_fail_reason.append('AC diff')
                    plate_report_ws['B'+ str(Paste_row+2)] = 'Re-Run Plate'
                    plate_report_ws['D'+ str(Paste_row+6)] = 'AC Diff Fail'
                    Plate_Fail = True
                    


                    
        #Samples needing re-run from Plate Failiures
        if len(plate_fail_reason) > 0:
            for i in range(0,len(SPK_Ct)): #First sample failiure error added
                Re_Run_Sample_Info_Dataframe = Re_Run_Sample_Info_Dataframe.append({'Need_Re_Run_Plate':data_file.split('\\')[-1][:-4],'Need_Re_Run_Samples':SPK_Ct['Sample Name'][i].split('+')[0],'Reason_Re_Run': ', '.join(plate_fail_reason) + ' Failed'}, ignore_index=True)



                   
        #Samples needing re-run from Replicate Difference 

        replicate_diff_issue_samples = []

        
        for Replicate_value_index in range(0,len(Replicate_Differences)):
            
            multiple_re_run_sample_reason = False #check if sample had other reasons to fail
            

            if str(Replicate_Differences['Ct diff'][Replicate_value_index]).startswith('PROBLEM'):
                    
                re_run_status = 'red'
                Samples_Failed = True

                    
                if Replicate_Differences['Sample Name'][Replicate_value_index].startswith('N') == False and Replicate_Differences['Sample Name'][Replicate_value_index].startswith('S') == False and Replicate_Differences['Sample Name'][Replicate_value_index].startswith('A') == False: #Name req for replicate diff

                    re_run_samples_replicates = 'true'

                    replicate_diff_issue_samples.append(Replicate_Differences['Sample Name'][Replicate_value_index])
                                       

                    #Update Re_Run_Samples with replicate diff fails, No STD rep diffs
                    for z in range(len(Re_Run_Sample_Info_Dataframe)): #is it in the re run sample df?
                        if Re_Run_Sample_Info_Dataframe['Need_Re_Run_Samples'][z] == Replicate_Differences['Sample Name'][Replicate_value_index] and data_file.split('\\')[-1][:-4] not in set(Re_Run_Sample_Info_Dataframe['Re_Run_Plate']): #its in the df

                            multiple_re_run_sample_reason = True
                            Re_Run_Sample_Info_Dataframe.iloc[z, 2] = Re_Run_Sample_Info_Dataframe.iloc[z, 2] + ' and Replicate Difference'
                            
                        
                if multiple_re_run_sample_reason == False and Replicate_Differences['Sample Name'][Replicate_value_index].startswith('N') == False and Replicate_Differences['Sample Name'][Replicate_value_index].startswith('S') == False and Replicate_Differences['Sample Name'][Replicate_value_index].startswith('A') == False: #if the sample starts with problem and no other fail reason
                    Re_Run_Sample_Info_Dataframe = Re_Run_Sample_Info_Dataframe.append({'Need_Re_Run_Plate':data_file.split('\\')[-1][:-4],'Need_Re_Run_Samples': Replicate_Differences['Sample Name'][Replicate_value_index].split('+')[0],'Reason_Re_Run':'Replicate Difference'}, ignore_index=True)


        plate_report_ws['C' + str(Paste_row+7)] = ', '.join(replicate_diff_issue_samples) #paste into 1 cell

        if len(replicate_diff_issue_samples) > 0:
            plate_report_ws['D'+ str(Paste_row+7)] = 'Replicate Difference'



        #Samples needing re-run from Inhibition

        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
        inhibited_samples = []
        
        for Inhibited_Sample_Index in range(0,len(SPK_Ct['Inhibition?'])):

            multiple_re_run_sample_reason = False   #check if sample had other reasons to fail


            if SPK_Ct['Inhibition?'][Inhibited_Sample_Index] == 'True':

                plate_report_ws['C' + str(Paste_row+8)] = SPK_Ct['Sample Name'][Inhibited_Sample_Index]
                plate_report_ws['D'+ str(Paste_row+8)] = 'Inhibited'

                if SPK_Ct['Sample Name'][Inhibited_Sample_Index].startswith('N') == False and SPK_Ct['Sample Name'][Inhibited_Sample_Index].startswith('S') == False and SPK_Ct['Sample Name'][Inhibited_Sample_Index].startswith('A') == False: 
                    inhibited_samples.append(SPK_Ct['Inhibition?'][Inhibited_Sample_Index])
                
                re_run_samples_inhib = 'true'
                re_run_status = 'red'
                Samples_Failed = True
                Paste_row += 1


                #Update Re_Run_Samples with inhibited

                for z in range(len(Re_Run_Sample_Info_Dataframe['Need_Re_Run_Samples'])): #is it in the re run sample df?
                    if Re_Run_Sample_Info_Dataframe['Need_Re_Run_Samples'][z] == SPK_Ct['Sample Name'][Inhibited_Sample_Index].split('+')[0] and data_file.split('\\')[-1][:-4] not in set(Re_Run_Sample_Info_Dataframe['Re_Run_Plate']): #it is in the df

                        Re_Run_Sample_Info_Dataframe.iloc[z, 2] = Re_Run_Sample_Info_Dataframe.iloc[z, 2] + ' and Inhibition'
                        multiple_re_run_sample_reason = True
                        
                if multiple_re_run_sample_reason == False and SPK_Ct['Sample Name'][Inhibited_Sample_Index].startswith('N') == False and SPK_Ct['Sample Name'][Inhibited_Sample_Index].startswith('S') == False: #if the sample starts with problem and no other fail reason
                    Re_Run_Sample_Info_Dataframe = Re_Run_Sample_Info_Dataframe.append({'Need_Re_Run_Plate':data_file.split('\\')[-1][:-4],'Need_Re_Run_Samples':SPK_Ct['Sample Name'][Inhibited_Sample_Index].split('+')[0],'Reason_Re_Run':'Inhibited'}, ignore_index=True)

   
        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
        

        #Some Re Run Status Reasons
        
        if re_run_samples_replicates == 'true':
                plate_report_ws['B'+ str(Paste_row+7)] = 'Re-run Samples'

        if re_run_samples_inhib == 'true' and re_run_samples_replicates == 'not yet':
                plate_report_ws['B'+ str(Paste_row+8)] = 'Re-run Samples'

        if re_run_status == 'red':
                plate_report_ws['B'+ str(Paste_row+1)].fill = PatternFill(start_color='00FF0000',end_color='00FF0000',fill_type='solid')

        for x in Sample_Name_List[-9:]:
                if x.startswith('S') == False:
                        plate_report_ws['D' + str(Paste_row+1)] = 'Re-Run with a STD removed'



#######


        #Is Qty Mean within STD curve range? Did Slope Fail? Which STD to kick? + Reruns
        
        STD_DF = STD_info_DF(form10)

        
        if 'Slope Requirement' in plate_fail_reason or 'R2 Requirement' in plate_fail_reason:
            plate_report_ws['C' + str(Paste_row+1)] = STD_DF['statement'][0]
            
        
        below_std_curve_range_list = []
        above_std_curve_range_list = []
        min_STD_qty = 25
        max_STD_qty = 10000000            



        min_STD_qty = min(STD_DF['std_range'])
        max_STD_qty = max(STD_DF['std_range'])


        Paste_row = file_count * 40 + 1 #Reset Paste_Row position
        
        for i in range(len(QuantityMeans['Qty Mean'])):
            
            multiple_re_run_sample_reason = False

            if QuantityMeans['Qty Mean'][i] < min_STD_qty: #Report dont retest
                plate_report_ws['B' + str(Paste_row+32)] = 'Fine, Just Reporting'
                plate_report_ws['D' + str(Paste_row+32)] = 'Qty Mean below ' + str(min_STD_qty)

                if QuantityMeans['Sample Name'][i] not in below_std_curve_range_list:
                    below_std_curve_range_list.append(QuantityMeans['Sample Name'][i]) 
                    
                
            if QuantityMeans['Qty Mean'][i] > max_STD_qty: #Dilute, Rerun
                plate_report_ws['B' + str(Paste_row+33)] = 'Re-run Samples'
                plate_report_ws['D' + str(Paste_row+33)] = 'Qty Mean above ' + str(max_STD_qty)

                if QuantityMeans['Sample Name'][i] not in above_std_curve_range_list:
                    above_std_curve_range_list.append(QuantityMeans['Sample Name'][i])

                Samples_Failed = True



                #Update Re-Run Samples: above STD curve range
                
                for z in range(len(Re_Run_Sample_Info_Dataframe)): #is sample in re run df already?
                    if Re_Run_Sample_Info_Dataframe['Need_Re_Run_Samples'][z] == QuantityMeans['Sample Name'][i] and data_file.split('\\')[-1][:-4] not in set(Re_Run_Sample_Info_Dataframe['Re_Run_Plate']): #it is in the df
                        
                        Re_Run_Sample_Info_Dataframe.iloc[z, 2] = Re_Run_Sample_Info_Dataframe.iloc[z, 2] + ' and Qty Mean above STD curve range'
                        multiple_re_run_sample_reason = True
                            
                if multiple_re_run_sample_reason == False:
                    Re_Run_Sample_Info_Dataframe = Re_Run_Sample_Info_Dataframe.append({'Need_Re_Run_Plate':data_file.split('\\')[-1][:-4],'Need_Re_Run_Samples':QuantityMeans['Sample Name'][i],'Reason_Re_Run':'Qty Mean above STD curve range'}, ignore_index=True)

                    
        plate_report_ws['C' + str(Paste_row+32)] = ', '.join(below_std_curve_range_list)
        plate_report_ws['C' + str(Paste_row+33)] = ', '.join(above_std_curve_range_list)


        if len(nan_indexes) > 0: #Nan samples statement
            plate_report_ws['B' + str(Paste_row+34)] = 'NaN samples'
            plate_report_ws['C' + str(Paste_row+34)] = 'at idx ' + str(nan_indexes)
            plate_report_ws['D' + str(Paste_row+34)] = 'were ignored in calculations'

#######################################################################################################################################################################

#AC Report

        AC_report_df = AC_report(form10)
        AC_qty_list = []
        passing_plates_with_no_ac_count = 0
        

        #Add File Information to PlateInformation.txt!!! {AC_values, STD1_Ct, Y_Intercept} -- Trying to write everything in at the same time

        for i in range(1, len(form10)):
            if form10['Sample Name'][i] == form10['Sample Name'][i-1] and form10['Sample Name'][i] == 'STD1':
                STD1_Ct_avg = (float(form10['Ct'][i]) + float(form10['Ct'][i-1]))/2

        for i in range(1, len(form10)):
            if form10['Sample Name'][i] == form10['Sample Name'][i-1] and form10['Sample Name'][i] == 'STD4':
                STD4_Ct_avg = (float(form10['Ct'][i]) + float(form10['Ct'][i-1]))/2

        for i in range(1, len(form10)):
            if form10['Sample Name'][i] == form10['Sample Name'][i-1] and form10['Sample Name'][i] == 'STD8':
                STD8_Ct_avg = (float(form10['Ct'][i]) + float(form10['Ct'][i-1]))/2


                
        
        if AC_report_df['AC_Value_Exists'][0] == True and 'Y_Intercept' in form10.columns:

            ac1 = float(AC_report_df['Qty1'][0])
            ac2 = float(AC_report_df['Qty2'][0])
            y_int = float(form10['Y_Intercept'][0])

            Plate_Info_DF = Plate_Info_DF.append({'Plate':data_file.split('\\')[-1][:-4],'Date_Processed':time.asctime(),'File_Length':len(form10)+12,'Plate_Failed':Plate_Fail,'Samples_Failed':Samples_Failed,'AC_qty1':ac1,'AC_qty2':ac2,'Y_Intercept':y_int,'STD1_Ct_avg':STD1_Ct_avg,'STD4_Ct_avg':STD4_Ct_avg,'STD8_Ct_avg':STD8_Ct_avg},ignore_index=True)

            
#            with open(Project_Output_Path + '\\' + 'Plate_Information.txt', 'a', newline = '\r') as Plate_Information:
#                Plate_Information.write('\n'+data_file.split('\\')[-1][:-4]+'\t'+time.asctime()+'\t'+str(len(form10)+12)+'\t'+str(Plate_Fail)+'\t'+str(Samples_Failed)+'\t'+ac1+'\t'+ac2+'\t'+y_int+'\t'+str(STD1_Ct_avg)+'\t'+str(STD4_Ct_avg)+'\t'+str(STD8_Ct_avg))
#            Plate_Information.close()

        elif AC_report_df['AC_Value_Exists'][0] == False and 'Y_Intercept' in form10.columns:

            y_int = float(form10['Y_Intercept'][0])

            Plate_Info_DF = Plate_Info_DF.append({'Plate':data_file.split('\\')[-1][:-4],'Date_Processed':time.asctime(),'File_Length':len(form10)+12,'Plate_Failed':Plate_Fail,'Samples_Failed':Samples_Failed,'Y_Intercept':y_int,'STD1_Ct_avg':STD1_Ct_avg,'STD4_Ct_avg':STD4_Ct_avg,'STD8_Ct_avg':STD8_Ct_avg},ignore_index=True)

#            with open(Project_Output_Path + '\\' + 'Plate_Information.txt', 'a', newline = '\r') as Plate_Information:
#                Plate_Information.write('\n'+data_file.split('\\')[-1][:-4]+'\t'+time.asctime()+'\t'+str(len(form10)+12)+'\t'+str(Plate_Fail)+'\t'+str(Samples_Failed)+'\t\t\t'+y_int+'\t'+str(STD1_Ct_avg)+'\t'+str(STD4_Ct_avg)+'\t'+str(STD8_Ct_avg)) #No AC Values
#            Plate_Information.close()

        elif AC_report_df['AC_Value_Exists'][0] == True and 'Y_Intercept' in form10.columns == False:

            ac1 = float(AC_report_df['Qty1'][0])
            ac2 = float(AC_report_df['Qty2'][0])

            Plate_Info_DF = Plate_Info_DF.append({'Plate':data_file.split('\\')[-1][:-4],'Date_Processed':time.asctime(),'File_Length':len(form10)+12,'Plate_Failed':Plate_Fail,'Samples_Failed':Samples_Failed,'AC_qty1':ac1,'AC_qty2':ac2,'STD1_Ct_avg':STD1_Ct_avg,'STD4_Ct_avg':STD4_Ct_avg,'STD8_Ct_avg':STD8_Ct_avg},ignore_index=True)

#            with open(Project_Output_Path + '\\' + 'Plate_Information.txt', 'a', newline = '\r') as Plate_Information:
#                Plate_Information.write('\n'+data_file.split('\\')[-1][:-4]+'\t'+time.asctime()+'\t'+str(len(form10)+12)+'\t'+str(Plate_Fail)+'\t'+str(Samples_Failed)+'\t'+ac1+'\t'+ac2+'\t\t'+str(STD1_Ct_avg)+'\t'+str(STD4_Ct_avg)+'\t'+str(STD8_Ct_avg)) 
#            Plate_Information.close()

        elif AC_report_df['AC_Value_Exists'][0] == False and 'Y_Intercept' in form10.columns == False:

            Plate_Info_DF = Plate_Info_DF.append({'Plate':data_file.split('\\')[-1][:-4],'Date_Processed':time.asctime(),'File_Length':len(form10)+12,'Plate_Failed':Plate_Fail,'Samples_Failed':Samples_Failed,'STD1_Ct_avg':STD1_Ct_avg,'STD4_Ct_avg':STD4_Ct_avg,'STD8_Ct_avg':STD8_Ct_avg},ignore_index=True)

#            with open(Project_Output_Path + '\\' + 'Plate_Information.txt', 'a', newline = '\r') as Plate_Information:
#                Plate_Information.write('\n'+data_file.split('\\')[-1][:-4]+'\t'+time.asctime()+'\t'+str(len(form10)+12)+'\t'+str(Plate_Fail)+'\t'+str(Samples_Failed)+'\t\t\t\t'+str(STD1_Ct_avg)+'\t'+str(STD4_Ct_avg)+'\t'+str(STD8_Ct_avg)) #No AC Values
#            Plate_Information.close()
            
#        Plate_Info_DF = pd.read_csv(Project_Output_Path + '\\' + 'Plate_Information.txt', sep = '\t', lineterminator = '\r', header = 0)




        
        ac_report_ws['A1'] = 'Plate'
        ac_report_ws['B1'] = 'Sample Name'
        ac_report_ws['C1'] = 'Ct'
        ac_report_ws['D1'] = 'Qty'
        ac_report_ws['E1'] = 'Qty Mean'
        ac_report_ws['G1'] = 'Std Dev of Qty Mean'
        ac_report_ws['H1'] = 'Avg of Qty Mean'
        ac_report_ws['I1'] = 'Percent CV' #std/avg
        ac_report_ws['J1'] = '# Passing AC Plates'
        ac_report_ws['K1'] = '# Passing No AC Plates'
        ac_report_ws['L1'] = '# of Failed Plates'
        ac_report_ws['M1'] = '# Total Plates'


     
        
        if Plate_Fail == False:
            if AC_report_df['AC_Value_Exists'][0] == True: #Not every plate has AC
                
                ac_report_ws['A' + str(plates_passed * 2 + 2)] = data_file.split('\\')[-1][:-4]
                
                ac_report_ws['B' + str(plates_passed * 2 + 2)] = 'AC'
                ac_report_ws['B' + str(plates_passed * 2 + 3)] = 'AC'

                ac_report_ws['C' + str(plates_passed * 2 + 2)] = AC_report_df['Ct1'][0]
                ac_report_ws['C' + str(plates_passed * 2 + 3)] = AC_report_df['Ct2'][0]

                ac_report_ws['D' + str(plates_passed * 2 + 2)] = AC_report_df['Qty1'][0]
                ac_report_ws['D' + str(plates_passed * 2 + 3)] = AC_report_df['Qty2'][0]

                ac_report_ws['E' + str(plates_passed * 2 + 2)] = AC_report_df['QtyMean'][0]
               

                plates_passed += 1
            

        for i in Plate_Info_DF['AC_qty1']:
            if math.isnan(float(i)) == False:
                AC_qty_list.append(float(i))
            else:
                passing_plates_with_no_ac_count += 1              
            
        for i in Plate_Info_DF['AC_qty2']:
            if math.isnan(float(i)) == False:
                AC_qty_list.append(float(i))
            else:
                passing_plates_with_no_ac_count += 1 #Duplicate?


        if len(AC_qty_list) > 0:
            
            ac_report_ws['G2'] = statistics.stdev(AC_qty_list)
            ac_report_ws['H2'] = ((sum(AC_qty_list))/len(AC_qty_list))
            ac_report_ws['I2'] = (statistics.stdev(AC_qty_list) / ((sum(AC_qty_list))/len(AC_qty_list)))*100


        ac_report_ws['J2'] = list(Plate_Info_DF['Plate_Failed']).count(False) - (passing_plates_with_no_ac_count / 2)   
        ac_report_ws['K2'] = passing_plates_with_no_ac_count / 2
        ac_report_ws['L2'] = list(Plate_Info_DF['Plate_Failed']).count(True)
        ac_report_ws['M2'] = len([i for i in Plate_Info_DF['Date_Processed'] if str(i).startswith('Excluded') == False])



                                           
#######################################################################################################################################################################
                                
#Locate Area to write next set of Calcultions
        file_count += 1
        currentRow = (file_count + files_before) * 109 + 1



#Save Re Run Sample Info
        Re_Run_Sample_Info_Dataframe.to_csv(Project_Output_Path + '\\' + 'Re_Run_Sample_Info.txt', sep='\t', index = False, header = True, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')





########################################################################################################################################################################

#When the For each file loop is ended


    #Re Run Samples Report
        
    if files_before == 0:
        for r in dataframe_to_rows(Re_Run_Sample_Info_Dataframe.iloc[:,:-1], index=False, header=True):
            re_run_samples_ws.append(r)
    else:
        
        del wb['Form13_Re Run Samples']
        re_run_samples_ws = wb.create_sheet('Form13_Re Run Samples')
        
        for r in dataframe_to_rows(Re_Run_Sample_Info_Dataframe.iloc[:,:-1], index=False, header=True):
            re_run_samples_ws.append(r)



    #Y_Intercept Report

    del wb['Form15_Y Intercept Report']
    y_ws = wb.create_sheet('Form15_Y Intercept Report')


    mod_Plate_Info_DF = Plate_Info_DF.copy()
    
    drop_idx = []


    for i in range(len(mod_Plate_Info_DF)):                                 
        if mod_Plate_Info_DF['Date_Processed'][i].startswith('Excluded'):
            drop_idx.append(i)

    mod_Plate_Info_DF.drop(drop_idx, inplace=True)      #Safely process non-excluded plates

    Y_ints = pd.DataFrame()
    Y_ints_rounded_list = []
    Y_ints['Plate'] = mod_Plate_Info_DF['Plate']
    Y_ints['Y_Intercepts'] = mod_Plate_Info_DF['Y_Intercept']
        
    for Y in Y_ints['Y_Intercepts']:
        Y_ints_rounded_list.append(round(Y,0))
    
    Y_ints['Y_rounded'] = Y_ints_rounded_list
    Y_ints['STD1_Ct_avg'] = mod_Plate_Info_DF['STD1_Ct_avg']
    Y_ints['STD4_Ct_avg'] = mod_Plate_Info_DF['STD4_Ct_avg']
    Y_ints['STD8_Ct_avg'] = mod_Plate_Info_DF['STD8_Ct_avg']

    
    for y in dataframe_to_rows(Y_ints, index=False, header=True):
        y_ws.append(y)


        
    #Histogram
    plt.figure(1)
    try:
        x_range = list(range(int(min(Y_ints['Y_rounded'])-2), int(max(Y_ints['Y_rounded'])+3)))
    except:
        print('No Data found, are you sure the Data_Directory_Path contains the data?')
        xrange = 1

    fig = plt.hist(Y_ints_rounded_list, bins=x_range, align = 'left')
    plt.xlabel('Y-Intercept Values')
    plt.ylabel('Count')
    plt.title('Y-Intercept Values Distribution')
    plt.savefig(Project_Output_Path + '\\' + 'y_int.png')

    img = openpyxl.drawing.image.Image(Project_Output_Path + '\\' + 'y_int.png')
    img.anchor = 'D3'


    #Line Chart
    plt.figure(2)
    fig = plt.plot(Y_ints['Plate'], Y_ints['Y_Intercepts'], label = 'Y Intercepts')
    fig = plt.plot(Y_ints['Plate'], Y_ints['STD1_Ct_avg'], label = 'STD1 Ct avg')
    fig = plt.plot(Y_ints['Plate'], Y_ints['STD4_Ct_avg'], label = 'STD4 Ct avg')
    fig = plt.plot(Y_ints['Plate'], Y_ints['STD8_Ct_avg'], label = 'STD8 Ct avg')
    plt.xlabel('Plate')
    plt.ylabel('Value')
    plt.title('Y-Intercept and STD(1,4,8)-Ct Trend for this Project')
    plt.legend(bbox_to_anchor = (1,0.5), loc='best')
    plt.tight_layout()
    tix = list(range(1,len(Y_ints)+1))
    plt.xticks(ticks = tix, labels = [], fontsize ='xx-small')
    plt.savefig(Project_Output_Path + '\\' + 'y_trend.png')
    
    img2 = openpyxl.drawing.image.Image(Project_Output_Path + '\\' + 'y_trend.png')
    img2.anchor = 'K2'

    
    y_ws.add_image(img)
    y_ws.add_image(img2)


    y_ws['G1'] = 'Total Y_Int Values'
    y_ws['H1'] = 'Total Plates'

    y_ws['G2'] = len(Y_ints)
    y_ws['H2'] = len([i for i in Plate_Info_DF['Date_Processed'] if i.startswith('Excluded') == False])
    

    #Send Form16 to the back -- Keep worksheets in order
    myorder = [0,1,2,3,4,5,6,7,8,9,10,12,13,14,15,11]
    wb._sheets = [wb._sheets[i] for i in myorder]

    
        
#Set Column Width for Readability

    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['X'].width = 17

    plate_report_ws.column_dimensions['A'].width = 15
    plate_report_ws.column_dimensions['B'].width = 16
    plate_report_ws.column_dimensions['C'].width = 21
    plate_report_ws.column_dimensions['D'].width = 22
    plate_report_ws.column_dimensions['E'].width = 5
    plate_report_ws.column_dimensions['F'].width = 18
    #plate_report_ws.column_dimensions['G'].width = 8
    plate_report_ws.column_dimensions['H'].width = 17
    plate_report_ws.column_dimensions['I'].width = 11
    plate_report_ws.column_dimensions['J'].width = 13
    plate_report_ws.column_dimensions['K'].width = 10
    plate_report_ws.column_dimensions['L'].width = 5
    plate_report_ws.column_dimensions['M'].width = 10
    plate_report_ws.column_dimensions['N'].width = 10
    plate_report_ws.column_dimensions['O'].width = 12
    plate_report_ws.column_dimensions['P'].width = 5
    plate_report_ws.column_dimensions['Q'].width = 5

    re_run_samples_ws.column_dimensions['A'].width = 40
    re_run_samples_ws.column_dimensions['B'].width = 25
    re_run_samples_ws.column_dimensions['C'].width = 60
    re_run_samples_ws.column_dimensions['D'].width = 40
    re_run_samples_ws.column_dimensions['E'].width = 25

    ac_report_ws.column_dimensions['A'].width = 36
    ac_report_ws.column_dimensions['B'].width = 13
    ac_report_ws.column_dimensions['C'].width = 11
    ac_report_ws.column_dimensions['D'].width = 11
    ac_report_ws.column_dimensions['E'].width = 12
    ac_report_ws.column_dimensions['G'].width = 18
    ac_report_ws.column_dimensions['H'].width = 14
    ac_report_ws.column_dimensions['I'].width = 10
    ac_report_ws.column_dimensions['J'].width = 16
    ac_report_ws.column_dimensions['K'].width = 20
    ac_report_ws.column_dimensions['L'].width = 15
    ac_report_ws.column_dimensions['M'].width = 11

    y_ws.column_dimensions['A'].width = 11
    y_ws.column_dimensions['B'].width = 11
    y_ws.column_dimensions['C'].width = 10
    y_ws.column_dimensions['D'].width = 13
    y_ws.column_dimensions['E'].width = 13
    y_ws.column_dimensions['F'].width = 13
    y_ws.column_dimensions['G'].width = 16
    y_ws.column_dimensions['H'].width = 11



#Check if someone has SCS19 Open before saving

    #workbook_path = Project_Output_Path + '\\' + Excel_File_Name

    print('')
    print('')
    print('Saving...')

    if excel_file_exists == True:
        while True: 
            try:
                with open(workbook_path, 'r+'):            
                    break
            except IOError:
                input("The program cannot save to the SCS19 file while it is open. Please close it and Press Enter to retry")
                
#Save!


    wb.save(workbook_path)



#Save Back Up File

    if backup_file_exists == True:
        while True: 
            try:
                with open(Backup_Filename, 'r+'):            
                    break
            except IOError:
                input("The program cannot save to the Back up SCS19 file while it is open. Please close it and Press Enter to retry")
        
    if backup_file_exists == True:
        os.rename(Backup_Filename, Project_Output_Path + '\\' + 'Form10_BackUp_' + Excel_File_Name[:-5] + '_' + str(now.strftime('%m-%d-%Y_time_%H_%M')) + '.xlsm')
        shutil.copyfile(Project_Output_Path + '\\' + Excel_File_Name, Project_Output_Path + '\\' + 'Form10_BackUp_' + Excel_File_Name[:-5] + '_' + str(now.strftime('%m-%d-%Y_time_%H_%M')) + '.xlsm')
    else:
        shutil.copyfile(Project_Output_Path + '\\' + Excel_File_Name, Project_Output_Path + '\\' + 'Form10_BackUp_' + Excel_File_Name[:-5] + '_' + str(now.strftime('%m-%d-%Y_time_%H_%M')) + '.xlsm')




#Delete Y_Int Image
        
    if os.path.exists(Project_Output_Path + '\\' + 'y_int.png'):
        os.remove(Project_Output_Path + '\\' + 'y_int.png')

    if os.path.exists(Project_Output_Path + '\\' + 'y_trend.png'):
        os.remove(Project_Output_Path + '\\' + 'y_trend.png')
    



#Record Success within Files
        
    Plate_Info_DF.to_csv(Project_Output_Path + '\\' + 'Plate_Information.txt', sep='\t', index = False, header = True, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')        
    Plate_Info_DF = pd.read_csv(Project_Output_Path + '\\' + 'Plate_Information.txt', sep = '\t', header = 0)
        
    Re_Run_Sample_Info_Dataframe['Run_Success'] = True
    Plate_Info_DF['Run_Success'] = True

    Re_Run_Sample_Info_Dataframe.to_csv(Project_Output_Path + '\\' + 'Re_Run_Sample_Info.txt', sep='\t', index = False, header = True, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')
    Plate_Info_DF.to_csv(Project_Output_Path + '\\' + 'Plate_Information.txt', sep='\t', index = False, header = True, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')




########################################################################################################################################################################

    
#Run the Program


if execute_style == 'Python':
    
    try:
        execute(Project_ID, Data_Directory_Path, Excel_File_Name, Project_Output_Path, Excel_Template_Path, SlopeMin, SlopeMax, LOQ, LOD, R2, Inhib_Diff_Criteria, Repl_Diff_Criteria, Repl_Diff_Criteria_for_STD_1_to_7, Repl_Diff_Criteria_for_STD_8, SCS_Path_that_ran_program)


    except BaseException as e:  #Error Handling

        print('')
        print('######################################################################################')
        print('')
        print('Runtime Exception Occured, saving into error log file')
        print('')
        print('')
        
        import traceback

        exc_type, exc_value, exc_traceback = sys.exc_info()

        print(''.join(traceback.format_exception(exc_type, exc_value, exc_traceback)))

        print('#######################################################################################')
        print('')
        print('')
        print('Results were not saved into SCS19 because the run did not complete')

        
        #Run Log
        run_log('\n#######################################################################################\n\n' + time.asctime() + '\n\n')
        run_log(''.join(traceback.format_exception(exc_type, exc_value, exc_traceback)))
        run_log('\n\n\n\n#######################################################################################\n\n\n\n')
        run_log('Results were not saved into SCS19 because the run did not complete')

        #Audit Log
        shutil.copyfile(Project_Output_Path + '\\' + Project_ID + '_' + Runtime + '_AF10_run_log.txt', Data_Directory_Path + '\\' + 'Audit_Logs' + '\\' + Project_ID + '_' + Runtime + '_AF10_run_log.txt')



        Plate_Info_DF = pd.read_csv(Project_Output_Path + '\\' + 'Plate_Information.txt', sep = '\t', header = 0)
        Re_Run_Sample_Info_Dataframe = pd.read_csv(Project_Output_Path + '\\' + 'Re_Run_Sample_Info.txt', sep = '\t', header = 0)    


        for i in range(len(Re_Run_Sample_Info_Dataframe)):                  #Mystery: Upon Reprocessing Re-Run-Sample-Info file, the Run_Success column is filled with 1.0 instead of True
            if Re_Run_Sample_Info_Dataframe.iloc[i,5] == 1.0:
                Re_Run_Sample_Info_Dataframe.iloc[i,5] = True


        Plate_Info_DF = Plate_Info_DF[Plate_Info_DF.Run_Success == True]
        Re_Run_Sample_Info_Dataframe = Re_Run_Sample_Info_Dataframe[Re_Run_Sample_Info_Dataframe.Run_Success == True]

        
        Plate_Info_DF.to_csv(Project_Output_Path + '\\' + 'Plate_Information.txt', sep='\t', index = False, header = True, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')
        Re_Run_Sample_Info_Dataframe.to_csv(Project_Output_Path + '\\' + 'Re_Run_Sample_Info.txt', sep='\t', index = False, header = True, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')


        print('')
        input("Press Enter to close the program window")
        

        os._exit(0)
        

#####################################################################################################################################################################################################        

#Run the Program


if execute_style == 'VBA':

#CMD Line: python C:\AutoForm10.py "1" "2" "3" "4" "5" "6" "7" "8" "9"

    if __name__ == '__main__': #when command line calls our python file, do:

        
        try:
            execute(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6], sys.argv[7], sys.argv[8], sys.argv[9], sys.argv[10], sys.argv[11], sys.argv[12], sys.argv[13], sys.argv[14], sys.argv[15])

            
        except BaseException as e: #Error Handling
            
            print('')
            print('######################################################################################')
            print('')
            print('Runtime Exception Occured, saving into error log file')
            print('')
            print('')
            
            import traceback

            exc_type, exc_value, exc_traceback = sys.exc_info()

            print(''.join(traceback.format_exception(exc_type, exc_value, exc_traceback)))

            print('#######################################################################################')
            print('')
            print('')
            print('Results were not saved into SCS19 because the run did not complete')
            
            #Run Log
            run_log('\n#######################################################################################\n\n' + time.asctime() + '\n\n')
            run_log(''.join(traceback.format_exception(exc_type, exc_value, exc_traceback)))
            run_log('\n\n\n\n#######################################################################################\n\n\n\n')
            run_log('Results were not saved into SCS19 because the run did not complete')

            #Audit Log
            shutil.copyfile(Project_Output_Path + '\\' + Project_ID + '_' + Runtime + '_AF10_run_log.txt', Data_Directory_Path + '\\' + 'Audit_Logs' + '\\' + Project_ID + '_' + Runtime + '_AF10_run_log.txt')


            Plate_Info_DF = pd.read_csv(Project_Output_Path + '\\' + 'Plate_Information.txt', sep = '\t', header = 0)
            Re_Run_Sample_Info_Dataframe = pd.read_csv(Project_Output_Path + '\\' + 'Re_Run_Sample_Info.txt', sep = '\t', header = 0)    


            for i in range(len(Re_Run_Sample_Info_Dataframe)):                  #Mystery: Upon Reprocessing Re-Run-Sample-Info file, the Run_Success column is filled with 1.0 instead of True
                if Re_Run_Sample_Info_Dataframe.iloc[i,5] == 1.0:
                    Re_Run_Sample_Info_Dataframe.iloc[i,5] = True


            Plate_Info_DF = Plate_Info_DF[Plate_Info_DF.Run_Success == True]
            Re_Run_Sample_Info_Dataframe = Re_Run_Sample_Info_Dataframe[Re_Run_Sample_Info_Dataframe.Run_Success == True]

            
            Plate_Info_DF.to_csv(Project_Output_Path + '\\' + 'Plate_Information.txt', sep='\t', index = False, header = True, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')
            Re_Run_Sample_Info_Dataframe.to_csv(Project_Output_Path + '\\' + 'Re_Run_Sample_Info.txt', sep='\t', index = False, header = True, quoting=csv.QUOTE_NONE, quotechar='',  escapechar='$')


            print('')
            input("Press Enter to close the program window")

            os._exit(0)
            



#Wrapping Things Up

print('')
print(time.asctime())

run_log('\n\n######################################################################################\n\n\n\nRun Successful: ' + time.asctime())
shutil.copyfile(Project_Output_Path + '\\' + Project_ID + '_' + Runtime + '_AF10_run_log.txt', Data_Directory_Path + '\\' + 'Audit_Logs' + '\\' + Project_ID + '_' + Runtime + '_AF10_run_log.txt')


print('Done') #yippee

print('')
print('')
input("Press Enter to close the program window")
